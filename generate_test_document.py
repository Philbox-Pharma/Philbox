"""
Philbox - Comprehensive Software Testing Document Generator
Generates a professional .xlsx file with 9 sheets covering all testing types.
Expanded to 200+ test cases covering every module and feature.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
OUTPUT_PATH = r"d:\1.STUDY\FYP\Main\Philbox\Philbox_Software_Testing_Document.xlsx"

HEADERS = [
    "Test Case ID", "Module / Feature", "Test Case Title", "Test Description",
    "Pre-conditions", "Test Steps", "Test Data", "Expected Result",
    "Actual Result", "Status", "Priority", "Tested By", "Date"
]

COL_WIDTHS = [14, 22, 30, 40, 30, 50, 28, 40, 18, 14, 12, 14, 14]

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# ─── SHEET THEMES (header_fill_hex, tab_color_hex, alt_row_hex) ──────────────
THEMES = {
    "Unit Testing":          ("1B4F72", "1B4F72", "EBF5FB"),
    "Integration Testing":   ("1A5276", "1A5276", "D6EAF8"),
    "System Testing":        ("0E6655", "0E6655", "D1F2EB"),
    "Acceptance Testing":    ("7D6608", "7D6608", "FEF9E7"),
    "Performance Testing":   ("6C3483", "6C3483", "F4ECF7"),
    "Security Testing":      ("922B21", "922B21", "FDEDEC"),
    "Usability Testing":     ("1F618D", "1F618D", "EAF2F8"),
    "Compatibility Testing": ("1E8449", "1E8449", "EAFAF1"),
}

# ═══════════════════════════════════════════════════════════════════════════════
#  1. UNIT TESTING (30 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

unit_tests = [
    # ── Customer Auth (React) ──
    ["UT-001", "Customer Auth (React)", "Register form renders all fields",
     "Verify Register.jsx renders Full Name, Email, Password, Confirm Password, Contact Number, Gender, DOB fields and Create Account button",
     "Dev server running on localhost:5173",
     "1. Navigate to /register\n2. Inspect DOM for all expected input elements\n3. Verify 'Create Account' button exists",
     "N/A",
     "All 7 form fields and submit button render without errors", "", "", "High", "", ""],

    ["UT-002", "Customer Auth (React)", "Empty form submission shows validation errors",
     "Submit the registration form without filling any fields and verify per-field red error messages appear",
     "On /register page",
     "1. Leave all fields empty\n2. Click 'Create Account'\n3. Verify error messages appear below each required field",
     "All fields empty",
     "Error messages: 'Full Name is required', 'Email is required', 'Password is required', 'Please confirm your password'", "", "", "High", "", ""],

    ["UT-003", "Customer Auth (React)", "Password regex validation matches backend Joi schema",
     "Enter a password with special characters and verify frontend rejects it with correct message matching backend regex ^[a-zA-Z0-9]{3,30}$",
     "On /register page",
     "1. Enter Full Name: 'Test User'\n2. Enter Email: 'test@test.com'\n3. Enter Password: 'Pass@123!'\n4. Click submit\n5. Verify password error appears",
     "Password: Pass@123!",
     "Error: 'Password must be 3-30 characters, letters and numbers only'", "", "", "High", "", ""],

    ["UT-004", "Customer Auth (React)", "Contact number accepts only digits",
     "Enter alphabetic characters in contact number field and verify validation rejects it",
     "On /register page",
     "1. Fill all required fields correctly\n2. Enter 'abc12345' in Contact Number\n3. Submit form\n4. Verify contact number error",
     "Contact: abc12345",
     "Error: 'Contact number must contain digits only'", "", "", "Medium", "", ""],

    ["UT-005", "Customer Auth (React)", "Password and Confirm Password mismatch error",
     "Enter different values in Password and Confirm Password fields and verify mismatch error is shown",
     "On /register page",
     "1. Enter Password: 'Test123'\n2. Enter Confirm Password: 'Test456'\n3. Click 'Create Account'\n4. Verify error below confirm password field",
     "Password: Test123, Confirm: Test456",
     "Error: 'Passwords do not match'", "", "", "High", "", ""],

    ["UT-006", "Customer Auth (React)", "Email format validation rejects invalid emails",
     "Enter various invalid email formats and verify rejection with meaningful error",
     "On /register page",
     "1. Enter email: 'notanemail'\n2. Submit -> verify error\n3. Enter email: 'user@'\n4. Submit -> verify error\n5. Enter email: '@domain.com'\n6. Submit -> verify error",
     "Emails: 'notanemail', 'user@', '@domain.com'",
     "Error: 'Please enter a valid email address' for each invalid format", "", "", "High", "", ""],

    # ── Backend Auth Controllers ──
    ["UT-007", "Admin Auth Controller (Node.js)", "Login controller returns session on valid credentials",
     "Unit test the admin login controller function with valid email/password to verify it returns session cookie",
     "Valid admin account exists in DB",
     "1. Call login controller with mock req/res\n2. Pass valid email and password\n3. Verify session is set on req.session\n4. Verify response status is 200",
     "Email: admin@philbox.com, Pass: Admin123",
     "Response: { status: 200, message: 'Login successful', data: { admin: {...} } }", "", "", "High", "", ""],

    ["UT-008", "Admin Auth Controller (Node.js)", "Login controller rejects invalid password",
     "Unit test the admin login controller rejects a wrong password and returns 401 without leaking info",
     "Valid admin account exists in DB",
     "1. Call login controller with valid email but wrong password\n2. Verify response status is 401\n3. Verify error message is generic 'Invalid credentials'\n4. Verify no password hash or user data leaked",
     "Email: admin@philbox.com, Pass: wrongpass",
     "Response: { status: 401, message: 'Invalid credentials' }, no sensitive data in response", "", "", "High", "", ""],

    # ── MongoDB Models ──
    ["UT-009", "Customer Model (MongoDB)", "Customer schema rejects duplicate emails",
     "Attempt to create two Customer documents with the same email and verify MongoDB throws duplicate key error",
     "Empty Customer collection",
     "1. Create Customer with email test@test.com\n2. Save successfully\n3. Try creating another Customer with same email\n4. Catch and verify error",
     "Email: test@test.com (duplicate)",
     "Mongoose throws 'E11000 duplicate key error' on second save", "", "", "High", "", ""],

    ["UT-010", "Customer Model (MongoDB)", "Customer schema makes passwordHash optional for OAuth users",
     "Verify that a Customer document with oauthId can be saved without passwordHash as per the conditional required validator",
     "Empty Customer collection",
     "1. Create Customer with oauthId: 'google-12345' and NO passwordHash\n2. Call .save()\n3. Verify document saves successfully\n4. Create another Customer WITHOUT oauthId and WITHOUT passwordHash\n5. Verify validation error on second save",
     "OAuth user without password vs regular user without password",
     "OAuth user saves OK; non-OAuth user fails with 'passwordHash is required'", "", "", "Medium", "", ""],

    ["UT-011", "Medicine Model (MongoDB)", "Medicine mgs field validates dosage format",
     "Verify the Medicine schema mgs validator accepts '500mg', '120ml' but rejects 'abc' and '500kg'",
     "N/A",
     "1. Create Medicine with mgs: '500mg' -> verify save OK\n2. Create Medicine with mgs: '120ml' -> verify save OK\n3. Create Medicine with mgs: 'abc' -> verify validation error\n4. Create Medicine with mgs: '500kg' -> verify validation error",
     "mgs values: '500mg', '120ml', 'abc', '500kg'",
     "'500mg' and '120ml' pass validation; 'abc' and '500kg' fail with message", "", "", "Medium", "", ""],

    ["UT-012", "Medicine Model (MongoDB)", "Medicine prescription_required auto-set for narcotics category",
     "Verify the pre-validate hook automatically sets prescription_required=true when category is 'Narcotics'",
     "MedicineCategory 'Narcotics' exists in DB",
     "1. Create Medicine with category referencing 'Narcotics'\n2. Call .save()\n3. Verify prescription_required is automatically set to true\n4. Create Medicine with category 'Pain Relief'\n5. Verify prescription_required remains false",
     "Categories: Narcotics, Pain Relief",
     "Narcotics medicine: prescription_required=true; Pain Relief: prescription_required=false", "", "", "High", "", ""],

    ["UT-013", "Coupon Model (MongoDB)", "Coupon schema rejects expired dates on save",
     "Verify the pre-save hook rejects a coupon with expiry_time in the past",
     "N/A",
     "1. Create Coupon with expiry_time set to yesterday\n2. Call .save()\n3. Catch error\n4. Verify error message: 'Expiry date must be in the future'",
     "expiry_time: yesterday's date",
     "Mongoose throws 'Expiry date must be in the future'", "", "", "Medium", "", ""],

    ["UT-014", "Coupon Model (MongoDB)", "Coupon percent_off bounded between 0 and 100",
     "Verify the Coupon schema min/max validators reject percent_off outside 0-100 range",
     "N/A",
     "1. Create Coupon with percent_off: -10 -> verify error\n2. Create Coupon with percent_off: 150 -> verify error\n3. Create Coupon with percent_off: 20 -> verify save OK",
     "percent_off: -10, 150, 20",
     "-10 and 150 fail validation; 20 saves successfully", "", "", "Medium", "", ""],

    ["UT-015", "DoctorSlot Model (MongoDB)", "DoctorSlot slot_duration bounded 0-20 minutes",
     "Verify the DoctorSlot schema rejects slot_duration > 20 or < 0",
     "N/A",
     "1. Create DoctorSlot with slot_duration: 25 -> verify validation error\n2. Create DoctorSlot with slot_duration: -5 -> verify validation error\n3. Create DoctorSlot with slot_duration: 15 -> verify save OK",
     "slot_duration: 25, -5, 15",
     "25 and -5 fail; 15 saves successfully", "", "", "Medium", "", ""],

    ["UT-016", "RefillReminder Model (MongoDB)", "RefillReminder timeOfDay validates 24-hour format",
     "Verify the regex validator /^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$/ on timeOfDay field",
     "N/A",
     "1. Create RefillReminder with timeOfDay: '08:30' -> save OK\n2. timeOfDay: '23:59' -> save OK\n3. timeOfDay: '25:00' -> validation error\n4. timeOfDay: 'morning' -> validation error",
     "timeOfDay: '08:30', '23:59', '25:00', 'morning'",
     "'08:30' and '23:59' pass; '25:00' and 'morning' fail validation", "", "", "Low", "", ""],

    ["UT-017", "Order Model (MongoDB)", "Order status enum rejects invalid values",
     "Verify Order schema rejects status values not in enum [pending, processing, on-the-way, cancelled-by-customer, completed, refunded]",
     "Valid customer and branch exist",
     "1. Create Order with status: 'shipped' -> verify validation error\n2. Create Order with status: 'on-the-way' -> verify save OK",
     "status: 'shipped' (invalid), 'on-the-way' (valid)",
     "'shipped' fails validation; 'on-the-way' saves successfully", "", "", "Medium", "", ""],

    ["UT-018", "Transaction Model (MongoDB)", "Transaction payment_method validates enum values",
     "Verify Transaction rejects payment_method values not in [Stripe-Card, JazzCash-Wallet, EasyPaisa-Wallet]",
     "N/A",
     "1. Create Transaction with payment_method: 'PayPal' -> verify error\n2. Create with 'JazzCash-Wallet' -> verify save OK\n3. Create with 'EasyPaisa-Wallet' -> verify save OK",
     "payment_method: 'PayPal', 'JazzCash-Wallet', 'EasyPaisa-Wallet'",
     "'PayPal' fails; JazzCash and EasyPaisa save successfully", "", "", "Medium", "", ""],

    # ── Backend Services ──
    ["UT-019", "Cart Service (Node.js)", "getCartCount returns correct item count",
     "Unit test the cart service getCartCount method to verify it aggregates item quantities correctly",
     "Customer has 3 items with quantities 2, 1, 3 in cart",
     "1. Mock CartItem.find() to return 3 items\n2. Call getCartCount(customerId)\n3. Verify returned itemCount equals 6\n4. Verify uniqueItems equals 3",
     "3 cart items with qty 2+1+3",
     "Returns { itemCount: 6, uniqueItems: 3 }", "", "", "Medium", "", ""],

    ["UT-020", "Catalog Service (Node.js)", "browseMedicines applies filters correctly",
     "Unit test that the browseMedicines service constructs correct MongoDB query with category and brand filters",
     "Medicine collection has various categories",
     "1. Call browseMedicines({ category: 'Pain Relief', brand: 'GSK' })\n2. Verify internally built query includes both filters\n3. Verify returned data is paginated",
     "Category: Pain Relief, Brand: GSK",
     "Query includes { category: 'Pain Relief', brand: 'GSK' }, results paginated with totalPages", "", "", "Medium", "", ""],

    ["UT-021", "SalespersonTask Service (Node.js)", "updateTaskStatus rejects invalid status transitions",
     "Unit test that a task in 'completed' status cannot transition back to 'pending'",
     "Task exists with status 'completed'",
     "1. Mock SalespersonTask.findById returning status: 'completed'\n2. Call updateTaskStatus with newStatus: 'pending'\n3. Verify controller returns 400 error",
     "Current status: completed, New status: pending",
     "400 Bad Request: 'Cannot revert completed task to pending'", "", "", "Medium", "", ""],

    # ── Joi DTOs ──
    ["UT-022", "Customer Auth DTO (Joi)", "customerRegisterDTO rejects missing fullName",
     "Validate that the Joi schema for customer registration returns 'fullName is required' when fullName is omitted",
     "N/A",
     "1. Call customerRegisterDTO.validate({ email: 'a@b.com', password: 'abc123' })\n2. Check error.details\n3. Verify error references fullName field",
     "Missing fullName from payload",
     "Joi returns: '\"fullName\" is required'", "", "", "High", "", ""],

    ["UT-023", "Appointment DTO (Joi)", "appointmentDTO rejects missing consultation_reason",
     "Validate that appointment booking DTO requires consultation_reason when appointment_request is 'processing'",
     "N/A",
     "1. Validate appointment DTO with doctor_id and preferred_date but without consultation_reason\n2. Verify Joi error\n3. Add consultation_reason and re-validate\n4. Verify success",
     "Missing consultation_reason",
     "Joi returns required field error without consultation_reason; passes with it", "", "", "Medium", "", ""],

    ["UT-024", "Slot DTO (Joi)", "slotDTO validates start_time HH:mm format",
     "Verify the slot creation DTO rejects improperly formatted start_time values",
     "N/A",
     "1. Validate with start_time: '9am' -> verify error\n2. Validate with start_time: '09:00' -> verify success\n3. Validate with start_time: '25:00' -> verify error",
     "start_time: '9am', '09:00', '25:00'",
     "'9am' and '25:00' fail; '09:00' passes validation", "", "", "Medium", "", ""],

    # ── React Components ──
    ["UT-025", "Medicine Detail (React)", "MedicineDetail quantity selector bounds check",
     "Verify the quantity selector does not allow values below 1 or above available stock",
     "Medicine page with stock = 10",
     "1. Mock catalogService.getMedicine with stock: 10\n2. Render MedicineDetail\n3. Try decrementing quantity below 1\n4. Try incrementing above 10",
     "Stock: 10",
     "Quantity stays clamped between 1 and 10", "", "", "Medium", "", ""],

    ["UT-026", "Doctor Profile (React)", "Specialization array renders safely when empty",
     "Render DoctorProfile.jsx with specialization: [] and verify it doesn't crash with .join() error",
     "Doctor profile loaded from API",
     "1. Mock profile API returning specialization: []\n2. Render DoctorProfile component\n3. Verify component renders without errors\n4. Check specialization area shows empty/fallback text",
     "specialization: []",
     "Component renders gracefully with empty/fallback placeholder, no .join() crash", "", "", "Medium", "", ""],

    ["UT-027", "Header Component (React)", "Cart badge re-fetches on cartUpdated event",
     "Verify Header.jsx listens for window 'cartUpdated' event and calls cartService.getCartCount()",
     "User is authenticated, Header is rendered",
     "1. Render Header component\n2. Mock cartService.getCartCount to return 5\n3. Dispatch window event 'cartUpdated'\n4. Wait for re-render\n5. Verify badge shows '5'",
     "Mock cart count: 5",
     "Badge text updates to '5' after event dispatch", "", "", "High", "", ""],

    ["UT-028", "Admin Dashboard (React)", "Analytics cards render correct counts from API",
     "Verify AdminDashboard.jsx displays correct summary counts from API response",
     "Admin logged in, dashboard API returns mock data",
     "1. Mock GET /api/admin/dashboard returning { totalCustomers: 120, totalOrders: 45, totalRevenue: 55000 }\n2. Render AdminDashboard component\n3. Verify 120, 45, 55000 appear in respective cards",
     "Mock dashboard stats",
     "All 3 stat cards render correct numbers from API response", "", "", "High", "", ""],

    ["UT-029", "Prescription Component (React)", "PrescriptionItem renders medicine details correctly",
     "Verify the PrescriptionItem renders medicine name, dosage, frequency, and duration from props",
     "Prescription data loaded",
     "1. Render PrescriptionItem with { medicineName: 'Amoxicillin', dosage: '500mg', frequency: 'Twice daily', duration: '7 days' }\n2. Verify all four values displayed\n3. Check for proper label-value pairing",
     "Medicine: Amoxicillin, 500mg, Twice daily, 7 days",
     "All prescription details render correctly with proper formatting", "", "", "Medium", "", ""],

    ["UT-030", "Appointment Component (React)", "BookAppointment renders doctor list with specializations",
     "Verify BookAppointment.jsx renders available doctors with specialization badges and available slot counts",
     "Mock doctors API returns 3 approved doctors",
     "1. Mock GET /doctors returning 3 doctors with specializations\n2. Render BookAppointment\n3. Verify 3 doctor cards rendered\n4. Verify each card shows doctor name, specialization badges, and slot count\n5. Verify 'Book' button exists on each card",
     "3 mock doctors with different specializations",
     "All doctor cards render with correct specialization badges and slot counts", "", "", "Medium", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  2. INTEGRATION TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

integration_tests = [
    ["IT-001", "Customer Registration Flow", "Frontend register -> Backend API -> Database",
     "Verify complete registration flow: React form submits payload, Express processes it, MongoDB saves Customer document",
     "Backend server running, MongoDB connected",
     "1. Fill registration form with valid data\n2. Submit form\n3. Verify POST /api/customer/auth/register is called\n4. Verify payload matches DTO\n5. Check MongoDB for new Customer document\n6. Verify verification email sent",
     "fullName: Test User, email: test@test.com, password: Test123",
     "201 response, Customer document in DB, verification email dispatched", "", "", "High", "", ""],

    ["IT-002", "Customer Login -> Session", "Login API sets session cookie correctly",
     "Verify successful login sets req.session.customerId and returns cookie to browser",
     "Verified customer account exists",
     "1. POST /api/customer/auth/login with valid credentials\n2. Check response for Set-Cookie header\n3. Verify session contains customerId and role\n4. Use cookie to access protected route /api/customer/cart",
     "Valid email/password",
     "Session cookie set, subsequent requests authenticated via cookie", "", "", "High", "", ""],

    ["IT-003", "Medicine Catalog API -> React UI", "GET /api/customer/medicines renders in Medicines.jsx",
     "Verify React component calls API, backend queries MongoDB, response renders as medicine cards",
     "Medicine collection has at least 5 documents",
     "1. Navigate to /medicines\n2. Verify GET /api/customer/medicines is called\n3. Check Network response contains medicine array\n4. Verify medicine cards render with Name, price, image",
     "N/A",
     "Medicine cards match API response data; pagination controls work", "", "", "High", "", ""],

    ["IT-004", "Add to Cart API", "POST /api/customer/cart/items creates cart entry",
     "Verify clicking Add to Cart sends correct medicine_id and quantity, CartItem created in MongoDB",
     "User logged in, medicine exists",
     "1. Navigate to /medicines/:id\n2. Set quantity to 2\n3. Click 'Add to Cart'\n4. Verify POST sent with { medicine_id, quantity: 2 }\n5. Check MongoDB CartItem for new entry",
     "medicine_id: valid, quantity: 2",
     "201 response, CartItem created, header cart badge updates", "", "", "High", "", ""],

    ["IT-005", "Appointment Booking -> Doctor View", "POST /api/customer/appointments/requests creates request",
     "Verify appointment request creation from customer UI through backend to MongoDB",
     "Customer and Doctor exist, Doctor has available slots",
     "1. Navigate to /appointments/book\n2. Select doctor and time\n3. Submit booking request\n4. Verify POST fires\n5. Check MongoDB for new appointment with status 'processing'",
     "doctor_id, preferred_date, consultation_reason",
     "Appointment request created with status 'processing'", "", "", "High", "", ""],

    ["IT-006", "Cancel Appointment Flow", "Cancellation updates appointment status in DB",
     "Verify cancellation from React UI updates appointment status to 'cancelled' in MongoDB",
     "Customer has a pending appointment request",
     "1. Navigate to /appointments\n2. Click Cancel on pending request\n3. Enter cancellation reason\n4. Confirm cancellation\n5. Verify API call succeeds\n6. Refresh and verify status changed",
     "cancellation_reason: 'Schedule conflict'",
     "Appointment status updates to cancelled in DB", "", "", "High", "", ""],

    ["IT-007", "Auth Middleware -> Protected Routes", "Unauthenticated request returns 401",
     "Verify Express auth middleware blocks requests without valid session cookie",
     "No active session",
     "1. GET /api/customer/cart WITHOUT session cookie -> verify 401\n2. Same request WITH valid cookie -> verify 200",
     "No auth cookie",
     "401 without cookie, 200 with valid cookie", "", "", "High", "", ""],

    ["IT-008", "Cart Update -> Header Sync", "Cart CRUD triggers header badge update via window event",
     "Verify cart quantity update fires 'cartUpdated' event and Header re-fetches count",
     "Cart has items, user on /cart page",
     "1. Update quantity from 1 to 3\n2. Verify PATCH fires\n3. Verify 'cartUpdated' event dispatches\n4. Verify Header badge number changes",
     "Quantity change: 1 -> 3",
     "Header badge reflects updated cart total in real-time", "", "", "Medium", "", ""],

    ["IT-009", "Doctor Slot -> Appointment Availability", "Created slots appear in customer booking",
     "Verify doctor-created slot becomes available in customer appointment booking",
     "Doctor approved, no existing slots",
     "1. Login as Doctor -> create slot tomorrow 10:00-11:00\n2. Verify slot in MongoDB\n3. Login as Customer -> /appointments/book\n4. Select that doctor\n5. Verify 10:00-11:00 slot available",
     "Slot: tomorrow 10:00-11:00",
     "Customer can see and select doctor's newly created slot", "", "", "Medium", "", ""],

    ["IT-010", "Admin Task -> Salesperson Tasks", "Admin creates task -> Salesperson sees it",
     "Verify task assignment by admin appears in salesperson's task list with correct status",
     "Admin and Salesperson accounts exist",
     "1. Login as Admin -> /admin/tasks\n2. Assign task to salesperson\n3. Login as Salesperson -> /salesperson/tasks\n4. Verify new task appears with 'pending' status",
     "Task: 'Restock Branch A', assignee: SP001",
     "Task visible in salesperson portal with pending status", "", "", "Medium", "", ""],

    ["IT-011", "Doctor Registration -> Application -> Admin Review", "Full doctor onboarding integration",
     "Verify registration API creates Doctor document, application stores DoctorApplication, admin fetches it",
     "Backend running, MongoDB connected",
     "1. POST /api/doctor/auth/register\n2. Verify Doctor document with status 'pending'\n3. POST application with profile details\n4. Verify DoctorApplication created\n5. GET /api/admin/doctors/applications\n6. Verify application appears",
     "Doctor: Dr. Ahmed, license: LIC-12345",
     "Doctor + DoctorApplication documents created, visible in admin review queue", "", "", "High", "", ""],

    ["IT-012", "Checkout Flow -> Order + Transaction", "Cart checkout creates Order and Transaction documents",
     "Verify completing checkout creates both Order and Transaction in MongoDB",
     "Customer logged in with items in cart",
     "1. Navigate to /cart -> Checkout\n2. Fill shipping address\n3. Select payment method: JazzCash-Wallet\n4. Place order\n5. Check MongoDB Order document\n6. Check MongoDB Transaction with correct payment_method",
     "Cart with 2 items, payment: JazzCash-Wallet",
     "Order + Transaction created, correct payment_method and amounts, cart emptied", "", "", "High", "", ""],

    ["IT-013", "MongoDB Connection Failure Handling", "Backend returns 500 when MongoDB unreachable",
     "Verify Express handles MongoDB disconnection gracefully without crashing",
     "Backend running, ability to simulate DB disconnect",
     "1. Stop MongoDB service\n2. Send GET /api/customer/medicines\n3. Verify 500 error with generic message\n4. Restart MongoDB\n5. Verify server recovers",
     "MongoDB service stopped",
     "500 error with safe message, server stays alive and recovers", "", "", "High", "", ""],

    ["IT-014", "Prescription Flow Cross-Portal", "Doctor writes prescription -> Customer views it",
     "Verify doctor creates prescription via API, customer retrieves it in their portal",
     "Active appointment between Customer and Doctor",
     "1. Login as Doctor -> active consultation\n2. Add prescription items\n3. Submit prescription\n4. Verify PrescriptionGeneratedByDoctor in MongoDB\n5. Login as Customer -> /prescriptions\n6. Verify prescription appears",
     "Medicine: Panadol 500mg, Twice daily, 5 days",
     "Prescription created by doctor visible in customer portal with all details", "", "", "High", "", ""],

    ["IT-015", "Salesperson Inventory Upload", "CSV upload creates/updates StockInHand per branch",
     "Verify salesperson CSV upload creates/updates StockInHand documents for assigned branch",
     "Salesperson assigned to Branch A, CSV prepared",
     "1. Login as Salesperson -> /salesperson/inventory\n2. Select Branch A\n3. Upload CSV with 10 items\n4. Verify POST with FormData fires (multer middleware)\n5. Check MongoDB StockInHand for 10 entries\n6. Verify each entry has correct branch_id",
     "CSV with 10 medicines, Branch: Branch A",
     "10 StockInHand documents created/updated with correct branch reference", "", "", "Medium", "", ""],

    ["IT-016", "Google OAuth -> Session Creation", "OAuth callback creates session for user",
     "Verify Google OAuth callback creates/finds customer account and establishes session",
     "Google OAuth configured in .env",
     "1. Navigate to /login -> 'Continue with Google'\n2. Complete Google sign-in\n3. Verify callback hit\n4. Check Customer document (created or matched by email)\n5. Verify session cookie set\n6. Verify redirect to /dashboard",
     "Valid Google account",
     "Session created, customer document exists, redirected to dashboard", "", "", "Medium", "", ""],

    ["IT-017", "RBAC Middleware Chain", "RBAC middleware enforces role checks across portals",
     "Verify rbac.middleware.js correctly enforces role-based access for all portal APIs",
     "Accounts for all 4 roles exist",
     "1. Customer -> GET /api/admin/branches -> 403\n2. Salesperson -> POST /api/doctor/slots -> 403\n3. Doctor -> GET /api/salesperson/tasks -> 403\n4. Admin -> GET /api/admin/branches -> 200",
     "All 4 role credentials",
     "Each cross-portal request returns 403 with role-specific error message", "", "", "High", "", ""],

    ["IT-018", "Feedback -> Doctor Rating Update", "Customer feedback updates doctor average rating",
     "Verify customer feedback submission updates Doctor document's average rating",
     "Completed appointment exists",
     "1. Login as Customer -> submit feedback rating: 4\n2. Verify Feedback document in MongoDB\n3. Query Doctor document -> verify averageRating updated",
     "Rating: 4/5, Comment: 'Good doctor'",
     "Feedback saved, Doctor.averageRating recalculated", "", "", "Medium", "", ""],

    ["IT-019", "Coupon Application -> Order Discount", "Applying coupon code reduces order total",
     "Verify coupon application during checkout correctly applies discount and tracks usage in Coupon model",
     "Valid active coupon exists with percent_off: 20",
     "1. Add items to cart (total: Rs.1000)\n2. Navigate to checkout\n3. Apply coupon code 'SAVE20'\n4. Verify discount applied: Rs.200 off\n5. Place order\n6. Check Order.total_after_applying_coupon = 800\n7. Check Coupon.times_used incremented by 1",
     "Coupon: SAVE20, 20% off, Cart: Rs.1000",
     "Order total reduced to Rs.800, coupon times_used incremented", "", "", "Medium", "", ""],

    ["IT-020", "Admin Branch -> Salesperson Assignment", "Admin assigns salesperson to branch -> reflected in SP portal",
     "Verify admin's branch-salesperson assignment is reflected in salesperson's managed branches",
     "Admin logged in, unassigned salesperson exists",
     "1. Login as Admin -> /admin/branches\n2. Create/edit branch -> assign salesperson\n3. Verify Branch document updated with salesperson_id\n4. Login as Salesperson\n5. Verify branch appears in dropdown on /salesperson/inventory",
     "Branch: Islamabad Central, Salesperson: SP001",
     "Salesperson sees assigned branch in their portal inventory dropdown", "", "", "Medium", "", ""],

    ["IT-021", "Search History Tracking", "Customer search queries saved in SearchHistory collection",
     "Verify that when a customer searches for medicine, the query is stored in SearchHistory MongoDB collection",
     "Customer logged in, medicine catalog available",
     "1. Login as Customer\n2. Navigate to /medicines\n3. Search for 'Panadol'\n4. Verify GET /api/customer/medicines?search=Panadol fires\n5. Check MongoDB SearchHistory collection for new entry with query 'Panadol'\n6. Verify entry has correct customer_id reference",
     "Search: 'Panadol'",
     "SearchHistory document created with query and customer reference", "", "", "Low", "", ""],

    ["IT-022", "Refill Reminder CRUD", "Customer creates reminder -> persisted in MongoDB -> appears in UI",
     "Verify the full refill reminder lifecycle from creation through API to display",
     "Customer logged in, medicines exist",
     "1. Navigate to /reminders\n2. Click 'Add Reminder'\n3. Select medicine, frequency: 'daily', timeOfDay: '08:00'\n4. Submit\n5. Verify POST fires\n6. Check MongoDB RefillReminder document\n7. Verify reminder appears in reminder list",
     "Medicine: Panadol, Frequency: daily, Time: 08:00",
     "RefillReminder created in DB, displayed in customer's reminders page", "", "", "Medium", "", ""],

    ["IT-023", "Admin Activity Logging", "Admin actions logged in AdminActivityLog collection",
     "Verify that admin operations (approve doctor, create task) generate activity log entries",
     "Admin logged in, pending doctor application exists",
     "1. Login as Admin\n2. Approve a pending doctor application\n3. Check AdminActivityLog collection\n4. Verify log entry with action type, admin_id, timestamp\n5. Navigate to /admin/analytics/activity-logs\n6. Verify the action appears in UI",
     "Action: Approve doctor application",
     "Activity log entry created with correct action, admin reference, and timestamp", "", "", "Low", "", ""],

    ["IT-024", "Admin Coupon Management", "Admin creates coupon -> Customer can apply at checkout",
     "Verify end-to-end coupon lifecycle from admin creation to customer usage",
     "Admin logged in",
     "1. Login as Admin -> create coupon: code 'FLAT10', 10% off, for 'medicine'\n2. Verify Coupon document in MongoDB\n3. Login as Customer -> add items to cart\n4. At checkout, apply 'FLAT10'\n5. Verify 10% discount applied\n6. Complete order -> verify coupon tracked on order",
     "Coupon: FLAT10, 10% off, for: medicine",
     "Coupon created by admin, usable by customer, discount applied correctly", "", "", "Medium", "", ""],

    ["IT-025", "Expired Coupon Rejection", "Expired coupon rejected at checkout with clear error",
     "Verify that an expired coupon cannot be applied and shows appropriate error to customer",
     "Coupon with past expiry_time exists in DB",
     "1. Login as Customer -> add items to cart\n2. Navigate to checkout\n3. Enter expired coupon code\n4. Click 'Apply'\n5. Verify error message: 'Coupon has expired'\n6. Verify total remains unchanged\n7. Verify Coupon.times_used NOT incremented",
     "Expired coupon code",
     "Error: 'Coupon has expired', total unchanged, usage not counted", "", "", "Medium", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  3. SYSTEM TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

system_tests = [
    ["ST-001", "Customer E2E Registration", "Register -> verify email -> login -> dashboard",
     "Test complete customer onboarding from registration through email verification to dashboard",
     "Fresh email, mail server accessible",
     "1. Navigate to /register\n2. Fill all fields\n3. Submit -> verify success\n4. Check email for verification link\n5. Click link\n6. Login with credentials\n7. Verify redirect to /dashboard\n8. Verify welcome message with name",
     "fullName: Ali Hassan, email: ali@test.com",
     "User lands on /dashboard with personalized greeting and functional quick links", "", "", "High", "", ""],

    ["ST-002", "Customer Medicine Purchase", "Search -> Add to Cart -> Checkout -> Order",
     "Test complete medicine purchase from searching to order confirmation",
     "Customer logged in, medicines in catalog",
     "1. Navigate to /medicines\n2. Search 'Panadol'\n3. Click result -> set qty 2\n4. Add to Cart\n5. Navigate to /cart -> verify item\n6. Checkout -> fill shipping\n7. Place order",
     "Medicine: Panadol, Qty: 2",
     "Order placed, confirmation shown with correct items and total", "", "", "High", "", ""],

    ["ST-003", "Doctor Onboarding Complete Flow", "Register -> Verify -> Profile -> Application -> Approval",
     "Test entire doctor onboarding lifecycle from registration to admin approval",
     "Fresh doctor email, admin account available",
     "1. /doctor/register -> register\n2. Verify email\n3. Login -> /doctor/complete-profile\n4. Fill specialization, license, experience\n5. Submit application\n6. Check application-status: 'Pending'\n7. Login as Admin -> approve\n8. Login as Doctor -> verify dashboard access",
     "Doctor: Dr. Sara, Specialization: Cardiology",
     "Doctor gains full dashboard access after admin approval", "", "", "High", "", ""],

    ["ST-004", "Appointment Lifecycle", "Book -> Accept -> Consult -> Prescribe -> Complete",
     "Test complete appointment lifecycle across customer and doctor portals",
     "Customer account, approved doctor with slots",
     "1. Customer books appointment\n2. Doctor accepts\n3. Customer verifies status: 'accepted'\n4. Doctor starts consultation -> writes prescription\n5. Doctor marks completed\n6. Customer views prescription in /prescriptions",
     "Appointment with Dr. Sara",
     "Full lifecycle: booking -> acceptance -> consultation -> prescription delivery", "", "", "High", "", ""],

    ["ST-005", "Role-Based Access Control", "Each role restricted to own portal routes",
     "Test that each RBAC role can only access its designated routes",
     "One account per role exists",
     "1. Customer -> /admin/dashboard -> redirect/403\n2. Customer -> /doctor/appointments -> redirect/403\n3. Doctor -> /admin/staff -> redirect/403\n4. Salesperson -> /admin/dashboard -> redirect/403\n5. Admin -> /admin/dashboard -> 200",
     "Various role credentials",
     "Each role restricted to own portal routes only", "", "", "High", "", ""],

    ["ST-006", "Salesperson Inventory Workflow", "Task -> Upload inventory -> Low stock alert",
     "Test complete salesperson workflow from task to inventory management",
     "Salesperson assigned to Branch A with tasks",
     "1. View task on /salesperson/tasks\n2. Navigate to /salesperson/inventory\n3. Select Branch A -> upload CSV\n4. Verify table updates\n5. Navigate to /salesperson/lowStockAlerts\n6. Verify alerts for Branch A\n7. Mark task completed",
     "Branch: Branch A, CSV with 50 items",
     "Inventory uploaded, low stock alerts generated, task marked complete", "", "", "High", "", ""],

    ["ST-007", "Admin Branch Management", "Create -> Assign Salesperson -> View Statistics",
     "Test admin branch lifecycle management",
     "Admin logged in, salesperson exists",
     "1. /admin/branches -> create branch\n2. Assign salesperson\n3. Navigate to branch statistics\n4. Verify branch in stats\n5. Verify salesperson shown",
     "Branch: Islamabad Central, Salesperson: SP001",
     "Branch created with salesperson assignment reflected in statistics", "", "", "Medium", "", ""],

    ["ST-008", "Session Persistence & Expiry", "Session persists across refreshes, expires after timeout",
     "Verify session management handles persistence and timeout correctly",
     "Valid customer session active",
     "1. Login -> /dashboard -> verify data\n2. Refresh (F5) -> still authenticated\n3. New tab /cart -> authenticated\n4. Delete cookie -> /dashboard -> redirect to /login",
     "Valid session cookie",
     "Session persists across refreshes/tabs but correctly expires", "", "", "High", "", ""],

    ["ST-009", "Multi-User Cart Isolation", "Two customers have independent carts",
     "Verify cart data isolation between customers",
     "Two customer accounts",
     "1. Customer A adds Medicine X\n2. Logout -> login as Customer B\n3. Customer B adds Medicine Y\n4. Verify B cart has only Y\n5. Switch to A -> verify only X",
     "Customer A: Medicine X, Customer B: Medicine Y",
     "Each customer sees only their own cart items", "", "", "High", "", ""],

    ["ST-010", "Admin Analytics Dashboard", "All analytics modules load with real data",
     "Verify all analytics (Revenue, Appointments, Engagement, Orders, Salesperson Performance) render correctly",
     "Admin logged in, historical data exists",
     "1. Navigate analytics section\n2. Revenue Analytics -> charts render\n3. Appointment Analytics -> charts render\n4. UserEngagement -> charts render\n5. OrdersAnalytics -> charts render\n6. SalespersonPerformance -> charts render\n7. Check console for errors",
     "N/A",
     "All charts render with data, no console errors or white screens", "", "", "Medium", "", ""],

    ["ST-011", "Error Recovery - Network Loss During Checkout", "User recovers from network loss mid-checkout",
     "Simulate network disconnection during checkout and verify no duplicate orders",
     "Customer with cart items, DevTools available",
     "1. /cart -> Checkout\n2. Fill shipping\n3. DevTools: set Network Offline\n4. Click 'Place Order' -> verify error toast\n5. Re-enable network\n6. Place Order again\n7. Verify only ONE order in MongoDB",
     "Network: simulated offline during checkout",
     "Error on failure, single order on retry, no duplicates", "", "", "High", "", ""],

    ["ST-012", "Password Reset Complete Flow", "Forgot -> Email -> Reset -> Login with new password",
     "Test entire password recovery flow end-to-end",
     "Verified customer account, email server accessible",
     "1. /forgot-password -> enter email\n2. Check inbox for reset link\n3. Click link -> /reset-password\n4. Enter new password\n5. Submit -> success\n6. Login with NEW password -> works\n7. Login with OLD password -> rejected",
     "Old: OldPass1, New: NewPass1",
     "Password reset successful, new works, old rejected", "", "", "High", "", ""],

    ["ST-013", "Admin Doctor Rejection", "Admin rejects application -> Doctor sees rejection reason",
     "Test rejection branch of doctor onboarding",
     "Doctor with submitted application, admin account",
     "1. Admin rejects with reason: 'Invalid license'\n2. Doctor checks application-status\n3. Verify 'Rejected' with reason\n4. Verify dashboard access blocked",
     "Rejection reason: 'Invalid license number'",
     "Doctor sees 'Rejected' with reason, dashboard blocked", "", "", "Medium", "", ""],

    ["ST-014", "Concurrent Cart Operations", "Two tabs updating cart simultaneously",
     "Verify cart consistency across concurrent browser tabs",
     "Customer logged in, medicine in catalog",
     "1. Tab 1: add Medicine A\n2. Tab 2: add Medicine B\n3. Tab 1: /cart -> verify A and B\n4. Tab 1: remove A\n5. Tab 2: refresh -> verify only B",
     "Two browser tabs, Medicines A and B",
     "Cart consistent across tabs, no data loss", "", "", "Medium", "", ""],

    ["ST-015", "Notification System E2E", "Actions generate notifications -> User views and reads",
     "Test that actions generate notifications in /notifications",
     "Customer with recent activity",
     "1. Login -> /notifications\n2. Verify appointment notification exists\n3. Click notification -> redirects\n4. Return -> verify marked as read\n5. Verify unread badge decrements",
     "Recent appointment acceptance + order",
     "Notifications generated, clickable, read state persists", "", "", "Medium", "", ""],

    ["ST-016", "Salesperson Low Stock Alert Workflow", "Low stock -> Alert -> Admin task -> Restock",
     "Verify low stock detection, alert generation, and task creation chain",
     "Branch with medicine, stock = 2, threshold = 5",
     "1. Salesperson -> /lowStockAlerts\n2. Verify medicine with qty < threshold in alerts\n3. Admin -> /admin/tasks\n4. Create restock task\n5. Salesperson -> verify task appears",
     "Stock: 2, Threshold: 5",
     "Alert visible, admin tasks salesperson, task appears in SP portal", "", "", "Medium", "", ""],

    ["ST-017", "Multi-Portal Data Consistency", "Admin updates medicine -> Reflects in all portals",
     "Verify medicine updates reflect across all portals",
     "Medicine exists, all portal accounts available",
     "1. Admin: update Medicine X price Rs.100 -> Rs.150\n2. Customer: /medicines -> verify Rs.150\n3. Salesperson: /inventory -> verify updated\n4. Doctor: prescription form -> verify Rs.150",
     "Medicine X: Rs.100 -> Rs.150",
     "Updated price consistent across all 4 portals", "", "", "High", "", ""],

    ["ST-018", "Customer Profile Update Flow", "Customer edits profile -> Changes persist across pages",
     "Verify customer profile update including name, contact, address is saved and reflected in header and sidebar",
     "Customer logged in with existing profile data",
     "1. Navigate to /profile\n2. Edit fullName to 'Updated Name'\n3. Edit contactNumber to '03001234567'\n4. Save profile\n5. Verify success toast\n6. Refresh page -> verify changes persist\n7. Check header -> verify name updated",
     "fullName: Updated Name, contactNumber: 03001234567",
     "Profile saved, changes persist on refresh, header reflects new name", "", "", "Medium", "", ""],

    ["ST-019", "Doctor Recurring Slot Creation", "Doctor creates recurring weekly slots -> Multiple slots generated",
     "Verify that creating a recurring weekly slot generates individual DoctorSlot entries for each occurrence",
     "Approved doctor logged in",
     "1. Navigate to /doctor/slots\n2. Create slot: Mon, 09:00-11:00, recurring: weekly, days: [1,3,5] (Mon/Wed/Fri)\n3. Set end_date: 4 weeks from now\n4. Submit\n5. Verify MongoDB has 12 DoctorSlot entries (3 per week x 4 weeks)\n6. Verify each slot has is_recurring: true",
     "Recurring weekly on Mon/Wed/Fri, 4 weeks",
     "12 individual DoctorSlot documents generated for each occurrence", "", "", "Medium", "", ""],

    ["ST-020", "Order Status Tracking Flow", "Order status transitions: pending -> processing -> on-the-way -> completed",
     "Verify order status correctly transitions through all valid states",
     "Customer has placed order, admin account available",
     "1. Customer places order -> status: 'pending'\n2. Admin updates to 'processing'\n3. Admin updates to 'on-the-way'\n4. Admin updates to 'completed'\n5. Customer views order -> status shows 'Completed' with green badge\n6. Verify each transition creates activity log",
     "Order through all status transitions",
     "Smooth status transitions, customer sees each status update", "", "", "Medium", "", ""],

    ["ST-021", "Admin Staff CRUD Complete Cycle", "Add -> View -> Edit -> Deactivate salesperson",
     "Test complete salesperson staff management lifecycle",
     "Admin logged in with manage_staff permission",
     "1. /admin/staff/salespersons -> Add Salesperson\n2. Fill details -> save -> verify appears in list\n3. Click salesperson -> view details page\n4. Edit phone number -> save\n5. Verify change persists\n6. Set status to 'Inactive'\n7. Verify salesperson cannot login",
     "Salesperson: Bilal, phone: 03001234567",
     "Full CRUD works, deactivated salesperson login blocked", "", "", "High", "", ""],

    ["ST-022", "Customer Order Cancellation and Refund", "Customer cancels order -> Refund initiated",
     "Verify customer can cancel a pending order and refund transaction is created",
     "Customer has a pending order",
     "1. Navigate to /orders\n2. Click on pending order\n3. Click 'Cancel Order'\n4. Confirm cancellation\n5. Verify order status: 'cancelled-by-customer'\n6. Verify refund Transaction document created with type: 'refund'\n7. Verify order.refund_status updated",
     "Pending order to cancel",
     "Order cancelled, refund transaction created, statuses updated correctly", "", "", "High", "", ""],

    ["ST-023", "Admin Feedback & Complaints Review", "Admin views customer feedback and complaints analytics",
     "Verify admin can access and filter feedback/complaints data in the analytics module",
     "Multiple customer feedback entries and complaints in DB",
     "1. Login as Admin\n2. Navigate to /admin/analytics -> FeedbackComplaints\n3. Verify feedback statistics charts render\n4. Filter by date range\n5. Verify data updates based on filter\n6. Check complaint details are viewable",
     "10+ feedback entries, 5+ complaints",
     "Feedback/complaint analytics render correctly with working filters", "", "", "Medium", "", ""],

    ["ST-024", "Branch Performance Statistics", "Admin views branch-wise performance comparison",
     "Verify BranchStatistics.jsx loads and displays comparative analytics for all branches",
     "Admin logged in, 3+ branches with order/inventory data",
     "1. Navigate to /admin/branches/statistics\n2. Verify all branches listed with performance metrics\n3. Verify revenue comparison chart renders\n4. Verify inventory health indicators per branch\n5. Click on a branch -> navigate to BranchDetails",
     "3+ branches with varying performance",
     "Branch comparison data loads, charts render, navigation to details works", "", "", "Medium", "", ""],

    ["ST-025", "Customer Search History Persistence", "Search history saved and accessible to customer",
     "Verify customer search queries are persisted and can be viewed/cleared",
     "Customer logged in, previously searched for medicines",
     "1. Navigate to /medicines\n2. Search 'Panadol' -> verify results\n3. Search 'Aspirin' -> verify results\n4. Navigate to search history section\n5. Verify 'Panadol' and 'Aspirin' appear\n6. Clear search history\n7. Verify history is empty",
     "Searches: 'Panadol', 'Aspirin'",
     "Search history recorded, displayed, and clearable", "", "", "Low", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  4. ACCEPTANCE TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

acceptance_tests = [
    ["AT-001", "Customer Registration", "User can create account successfully",
     "Verify new user can register and receive verification email per business requirement",
     "Valid email, internet connection",
     "1. Open Philbox homepage\n2. Navigate to /register\n3. Fill all mandatory fields\n4. Click 'Create Account'\n5. Check email for verification link\n6. Click link -> verified",
     "fullName: Ahmad Khan, email: ahmad@email.com",
     "Account created, verification email received within 2 minutes", "", "", "High", "", ""],

    ["AT-002", "Medicine Search & Purchase", "Customer can find and buy medicines",
     "Verify primary business function: users can find and buy medicines",
     "Customer logged in, medicines in catalog",
     "1. Search medicine by name\n2. View details (price, description, availability)\n3. Add to cart\n4. Navigate to cart\n5. Proceed to checkout\n6. Complete purchase",
     "Search: 'Aspirin', Qty: 1",
     "Medicine found, cart updated, checkout completed with order confirmation", "", "", "High", "", ""],

    ["AT-003", "Appointment Booking", "Customer can book doctor appointment",
     "Verify users can book appointments with doctors per telemedicine requirement",
     "Customer logged in, approved doctors with slots",
     "1. Navigate to appointment booking\n2. Browse doctors\n3. Select doctor and time slot\n4. Provide consultation reason\n5. Submit request\n6. Verify confirmation",
     "Doctor: Dr. Sara, Reason: Headache",
     "Request submitted, visible in 'My Appointments' with processing status", "", "", "High", "", ""],

    ["AT-004", "Doctor Availability Management", "Doctor can manage slot schedule",
     "Verify doctors can create, view, edit, and delete availability slots",
     "Approved doctor logged in",
     "1. Navigate to /doctor/slots\n2. Create new slot\n3. Verify slot on calendar\n4. Edit timing\n5. Delete slot\n6. Verify deleted slot gone",
     "New slot: Mon 09:00-10:00",
     "Doctor has full CRUD control over availability slots", "", "", "High", "", ""],

    ["AT-005", "Admin Staff Management", "Admin can manage staff (salespersons, co-admins)",
     "Verify admin can add, edit, and manage system users",
     "Admin logged in with full permissions",
     "1. /admin/staff/salespersons -> Add Salesperson\n2. Fill details -> save\n3. Edit phone number -> save\n4. Test Active/Inactive toggle",
     "New salesperson: Bilal, phone: 03001234567",
     "Salesperson CRUD operations work end-to-end", "", "", "High", "", ""],

    ["AT-006", "Inventory Management", "Salesperson can upload and manage inventory",
     "Verify salesperson can manage branch inventory through dedicated interface",
     "Salesperson logged in, assigned to branch",
     "1. /salesperson/inventory -> select branch\n2. View current items\n3. Upload new batch\n4. Verify items appear\n5. Check quantities match",
     "Branch: Lahore Central, CSV with 20 medicines",
     "Inventory updated, items visible, quantities matching", "", "", "High", "", ""],

    ["AT-007", "Responsive Design", "Application works on mobile and tablet",
     "Verify Philbox is usable across standard viewport sizes",
     "Application running on localhost",
     "1. Desktop (1920x1080) -> verify layout\n2. Tablet (768x1024) -> verify adapts\n3. Mobile (375x667) -> verify hamburger menu\n4. Navigate key pages on each\n5. No horizontal overflow",
     "Desktop, Tablet, Mobile viewports",
     "All pages render correctly across device sizes", "", "", "Medium", "", ""],

    ["AT-008", "Google OAuth Authentication", "User can sign in with Google",
     "Verify Google OAuth login flow works end-to-end",
     "Google OAuth configured",
     "1. /login -> 'Continue with Google'\n2. Google popup appears\n3. Select account\n4. Verify redirect to Philbox\n5. Session created, dashboard loads",
     "Valid Google account",
     "OAuth flow completes, session established, correct portal loads", "", "", "Medium", "", ""],

    ["AT-009", "Password Reset Flow", "User can reset forgotten password",
     "Verify forget/reset password flow works for customers and doctors",
     "Existing verified account",
     "1. /forgot-password -> enter email\n2. Send Reset Link\n3. Check email\n4. Click link -> /reset-password\n5. Enter new password\n6. Login with new password",
     "New password: NewPass123",
     "Password updated, new credentials work for login", "", "", "High", "", ""],

    ["AT-010", "Notification System", "Users receive relevant notifications",
     "Verify users see contextual notifications for platform actions",
     "User logged in with recent activity",
     "1. /notifications -> verify list populated\n2. Click notification -> navigates to relevant page\n3. Mark as read\n4. Verify read status persists",
     "N/A",
     "Notifications display, navigable, read state persists", "", "", "Low", "", ""],

    ["AT-011", "Customer Order History", "Customer can view past orders and track status",
     "Verify customers can access order history with status tracking",
     "Customer with 3+ past orders",
     "1. /orders -> verify list with dates, totals, status\n2. Click specific order -> verify item details\n3. Verify status badge (pending/processing/delivered)\n4. Verify pagination for 10+ orders",
     "3+ past orders",
     "Order history accurate with details, status tracking, pagination", "", "", "High", "", ""],

    ["AT-012", "Doctor Consultation Notes", "Doctor can write consultation notes and prescriptions",
     "Verify doctors can document consultation details during active appointments",
     "Doctor with active appointment",
     "1. /doctor/consultations -> open consultation\n2. Write diagnosis notes\n3. Add prescription items via CreatePrescriptionModal\n4. Save\n5. Verify data persists on refresh",
     "Diagnosis: 'Seasonal flu'",
     "Notes saved, prescription attached, persists across sessions", "", "", "High", "", ""],

    ["AT-013", "Admin Customer Management", "Admin can view and manage customer accounts",
     "Verify admin can search, view, and moderate customers",
     "Admin logged in, multiple customers exist",
     "1. /admin/customers -> list loads\n2. Search by name\n3. Click -> view CustomerDetails\n4. View order/appointment history\n5. Test account deactivation",
     "Customer search: 'Ahmad Khan'",
     "Admin can list, search, view details, manage accounts", "", "", "Medium", "", ""],

    ["AT-014", "Medicine Category Filtering", "Customer can filter medicines by category and manufacturer",
     "Verify catalog supports meaningful filtering",
     "Catalog with multiple categories",
     "1. /medicines -> select category: 'Pain Relief'\n2. Verify only pain relief shown\n3. Add manufacturer: 'GSK'\n4. Results narrow further\n5. Clear filters -> full catalog returns",
     "Category: Pain Relief, Manufacturer: GSK",
     "Filters narrow correctly, clearing restores full catalog", "", "", "Medium", "", ""],

    ["AT-015", "Salesperson Revenue Analytics", "Salesperson can view personal performance",
     "Verify salesperson can access and interpret personal revenue analytics",
     "Salesperson with sales history",
     "1. /salesperson/dashboard -> revenue analytics loads\n2. Total sales figure displayed\n3. Charts render with data\n4. Recent activity log shows latest actions",
     "Salesperson with 30+ days activity",
     "Revenue analytics display correctly, charts and figures match", "", "", "Medium", "", ""],

    ["AT-016", "Admin Roles & Permissions", "Admin can create roles with granular permissions",
     "Verify admin can configure custom roles with specific permissions",
     "Admin with super-admin permissions",
     "1. /admin/roles-permissions -> create 'Branch Manager' role\n2. Assign: manage_branches, view_inventory\n3. Deny: manage_staff, manage_doctors\n4. Save role\n5. Create admin with role\n6. Login as new admin -> verify only assigned permissions work",
     "Role: Branch Manager",
     "Custom role created, permissions enforced, restricted features inaccessible", "", "", "High", "", ""],

    ["AT-017", "Refill Reminder System", "Customer receives refill reminders",
     "Verify RefillReminder generates timely notifications for medicine refills",
     "Customer with completed prescription",
     "1. /reminders -> add reminder for medicine\n2. Select frequency: daily, time: 08:00\n3. Save -> verify reminder created\n4. Verify notification sent at scheduled time\n5. Edit reminder -> change to weekly\n6. Delete reminder -> verify removed",
     "Medicine: Amoxicillin, daily at 08:00",
     "Reminder CRUD works, notifications scheduled correctly", "", "", "Low", "", ""],

    ["AT-018", "Coupon and Discount", "Customer can apply coupon at checkout",
     "Verify coupon/discount system works during checkout",
     "Valid coupon exists, items in cart",
     "1. /cart -> Checkout\n2. Enter coupon: 'SAVE20'\n3. Apply -> discount shown\n4. Verify final total reflects discount\n5. Place order -> coupon reference on order",
     "Coupon: SAVE20, 20% off",
     "Coupon applied, discount reflected, order has coupon details", "", "", "Medium", "", ""],

    ["AT-019", "Appointment Type Selection", "Customer can choose between in-person and online appointments",
     "Verify both appointment types (in-person, online) are bookable with appropriate fields",
     "Approved doctor with both appointment types available",
     "1. Navigate to /appointments/book\n2. Select doctor\n3. Choose 'Online' type -> verify video call info field appears\n4. Choose 'In-person' -> verify location/branch info appears\n5. Complete booking with each type\n6. Verify appointment_type saved correctly in DB",
     "Appointment types: in-person, online",
     "Both types bookable, correct fields shown, type saved in DB", "", "", "Medium", "", ""],

    ["AT-020", "Admin Task Priority System", "Admin can assign tasks with priority levels",
     "Verify admin can create tasks with low/medium/high/urgent priority and salesperson sees correct priority",
     "Admin logged in, salesperson assigned to branch",
     "1. /admin/tasks -> create task with priority: 'urgent'\n2. Create another with priority: 'low'\n3. Verify both saved with correct priority\n4. Login as Salesperson\n5. Verify urgent task highlighted/ordered first\n6. Verify priority badges display correctly",
     "Tasks: urgent and low priority",
     "Priority levels set correctly, urgent shown prominently in SP portal", "", "", "Medium", "", ""],

    ["AT-021", "Doctor Patient Feedback View", "Doctor can see patient feedback and ratings",
     "Verify doctors can view patient feedback on completed consultations via PatientFeedback.jsx",
     "Approved doctor with completed consultations that have feedback",
     "1. Login as Doctor\n2. Navigate to /doctor/feedback\n3. Verify list of patient feedback entries\n4. Check each entry shows rating (1-5 stars), comment, and patient name\n5. Verify average rating calculation at top\n6. Verify pagination for multiple entries",
     "5+ feedback entries from patients",
     "Feedback list loads, ratings displayed, average calculated correctly", "", "", "Medium", "", ""],

    ["AT-022", "Prescription Required Medicine", "Medicines requiring prescription cannot be purchased without one",
     "Verify that narcotics/prescription-required medicines enforce prescription upload before adding to cart",
     "Customer logged in, narcotics medicine exists (prescription_required: true)",
     "1. Navigate to /medicines\n2. Find a prescription-required medicine\n3. Verify 'Prescription Required' badge visible\n4. Try to add to cart without prescription\n5. Verify error/prompt to upload prescription\n6. Upload prescription -> verify add to cart succeeds",
     "Narcotics medicine with prescription_required: true",
     "Prescription-required badge visible, cart blocked without prescription, allowed with", "", "", "High", "", ""],

    ["AT-023", "Admin Order Dashboard", "Admin can view and manage all orders from OrdersDashboard",
     "Verify the admin orders management interface shows all orders with filter/sort capabilities",
     "Admin logged in, 20+ orders in system",
     "1. Navigate to /admin/orders\n2. Verify all orders listed with customer name, total, status, date\n3. Filter by status: 'pending'\n4. Verify only pending orders shown\n5. Sort by date (newest first)\n6. Click order -> view details\n7. Update order status",
     "20+ orders with mixed statuses",
     "Orders dashboard loads, filters/sorts work, status updates save", "", "", "High", "", ""],

    ["AT-024", "Multiple Payment Methods", "Customer can pay using Stripe Card, JazzCash, or EasyPaisa",
     "Verify all 3 payment methods work correctly during checkout",
     "Customer with items in cart, payment gateways configured",
     "1. Checkout -> select 'Stripe-Card' -> verify card input fields appear\n2. Complete payment -> verify Transaction.payment_method = 'Stripe-Card'\n3. New order -> select 'JazzCash-Wallet' -> verify wallet number input\n4. Complete -> verify payment_method = 'JazzCash-Wallet'\n5. New order -> select 'EasyPaisa-Wallet'\n6. Verify payment_method = 'EasyPaisa-Wallet'",
     "3 payment methods: Stripe, JazzCash, EasyPaisa",
     "All 3 payment methods functional with correct UI and Transaction records", "", "", "High", "", ""],

    ["AT-025", "Admin Profile Management", "Admin can update own profile details",
     "Verify admin can edit their name, email, and password from the profile management page",
     "Admin logged in",
     "1. Navigate to /admin/profile\n2. Edit display name\n3. Save -> verify success\n4. Change password (old + new)\n5. Save -> verify success\n6. Logout -> login with new password\n7. Verify profile changes persist",
     "Name: Updated Admin, New password",
     "Profile updated, password changed, changes persist after re-login", "", "", "Medium", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  5. PERFORMANCE TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

performance_tests = [
    ["PT-001", "Page Load - Landing Page", "Landing page loads within 2 seconds",
     "Measure time from navigation start to DOMContentLoaded for customer landing page",
     "Dev server running, no throttling",
     "1. DevTools Performance tab\n2. Navigate to localhost:5173\n3. Record performance\n4. Measure DOMContentLoaded\n5. Verify LCP < 2.5s",
     "N/A",
     "DOMContentLoaded < 2.0s, LCP < 2.5s", "", "", "High", "", ""],

    ["PT-002", "Page Load - Dashboard", "Dashboard loads with 5 parallel API calls within 3 seconds",
     "Measure CustomerDashboard render time with 5 concurrent API requests",
     "Customer logged in with active data",
     "1. DevTools Network tab\n2. Navigate to /dashboard\n3. Verify 5 API calls fire simultaneously\n4. Measure time until all cards rendered",
     "N/A",
     "All 5 API responses received and UI fully rendered within 3 seconds", "", "", "High", "", ""],

    ["PT-003", "API Response - Medicine Catalog", "GET /api/customer/medicines < 500ms",
     "Measure backend response time for medicine catalog with pagination",
     "100+ medicines in database",
     "1. Use Postman\n2. GET /api/customer/medicines?page=1&limit=12\n3. Record time\n4. Repeat 10 times\n5. Calculate average",
     "page=1, limit=12",
     "Average response time < 500ms across 10 requests", "", "", "High", "", ""],

    ["PT-004", "API Response - Login", "POST /api/customer/auth/login < 800ms",
     "Measure login endpoint response time including bcrypt comparison",
     "Valid customer account exists",
     "1. POST with valid credentials\n2. Record time to 200 response\n3. Repeat 5 times\n4. Calculate average",
     "Valid email/password",
     "Average login response time < 800ms", "", "", "High", "", ""],

    ["PT-005", "MongoDB Query - Medicine Search", "Search query completes within 200ms",
     "Measure MongoDB query time for text-based medicine search",
     "500+ medicines indexed in MongoDB",
     "1. MongoDB Compass Explain Plan\n2. Execute: db.medicines.find({ Name: /panadol/i }).limit(12)\n3. Check execution time\n4. Verify query uses index",
     "Search: 'panadol'",
     "Query executes within 200ms using proper index", "", "", "Medium", "", ""],

    ["PT-006", "Concurrent Users", "Server handles 50 concurrent requests",
     "Stress test Express server with 50 simultaneous API requests",
     "Server running in production mode",
     "1. Use Artillery or JMeter\n2. 50 virtual users\n3. Each: GET /api/customer/medicines\n4. Run 30 seconds\n5. Measure avg response, error rate, throughput",
     "50 concurrent users, 30s",
     "Avg response < 2s, error rate < 1%, throughput > 20 req/s", "", "", "High", "", ""],

    ["PT-007", "React Bundle Size", "Main JS bundle under 500KB gzipped",
     "Verify Vite production build produces optimized chunks",
     "Source code ready for build",
     "1. npm run build\n2. Check dist/assets/\n3. Verify chunk sizes\n4. Check gzipped sizes\n5. Identify chunks > 500KB",
     "N/A",
     "No single chunk > 500KB gzipped; lazy loading on portal routes", "", "", "Medium", "", ""],

    ["PT-008", "Image Lazy Loading", "Below-fold images load only on scroll",
     "Verify medicine catalog images use lazy loading",
     "Page with multiple images",
     "1. DevTools Network -> filter 'img'\n2. Load /medicines\n3. Count initial image requests\n4. Scroll down\n5. Verify additional images load progressively",
     "N/A",
     "Only above-fold images load initially; below-fold on scroll", "", "", "Medium", "", ""],

    ["PT-009", "Slow Network Resilience (3G)", "App usable on Slow 3G",
     "Verify loading states under throttled conditions",
     "DevTools throttling available",
     "1. Network -> 'Slow 3G'\n2. /login -> form renders with loader\n3. /medicines -> skeleton appears\n4. Add to cart -> loading state\n5. No white screens",
     "Network: Slow 3G",
     "All actions show loading feedback; no white screens under 10s", "", "", "Medium", "", ""],

    ["PT-010", "Memory Leak Detection", "No memory leaks on page navigation",
     "Verify React components cleanup subscriptions on unmount",
     "Application running in browser",
     "1. DevTools Memory tab\n2. Take heap snapshot\n3. Navigate 10 pages rapidly\n4. Return to start\n5. Take another snapshot\n6. Compare for detached DOM nodes\n7. Check 'cartUpdated' listener cleanup",
     "N/A",
     "No significant detached DOM nodes; memory stabilizes", "", "", "Low", "", ""],

    ["PT-011", "Admin Dashboard with Analytics", "Admin dashboard loads within 4 seconds",
     "Measure admin dashboard load including analytics charts with real data",
     "Admin logged in, 6+ months data",
     "1. DevTools Performance\n2. Navigate /admin/dashboard\n3. Wait for all charts\n4. Record total time\n5. No chart shows 'Loading...' after 4s",
     "6 months analytics data",
     "Admin dashboard fully rendered with charts within 4 seconds", "", "", "High", "", ""],

    ["PT-012", "MongoDB Aggregation Pipeline", "Analytics aggregation < 1 second",
     "Measure MongoDB aggregation pipelines for revenue/appointment analytics",
     "Analytics collections populated",
     "1. MongoDB Compass\n2. Run revenue aggregation for 6-month range\n3. Record execution time\n4. Verify indexed $match stages",
     "Date range: 6 months",
     "Aggregation completes within 1 second using indexes", "", "", "Medium", "", ""],

    ["PT-013", "Concurrent Login Stress", "100 concurrent login requests handled",
     "Stress test login with 100 simultaneous bcrypt authentication requests",
     "100 valid accounts, load testing tool",
     "1. 100 virtual users hitting POST login simultaneously\n2. Measure avg and p95 response time\n3. Record error rate\n4. Monitor CPU/memory",
     "100 concurrent users",
     "Avg < 3s, p95 < 5s, error rate < 2%, server stable", "", "", "High", "", ""],

    ["PT-014", "Cart Operations Under Load", "50 users performing cart CRUD simultaneously",
     "Verify cart service handles concurrent operations without corruption",
     "50 test accounts with carts",
     "1. 50 users random cart operations\n2. Run 60 seconds\n3. Verify no 500 errors\n4. Check no orphaned CartItems\n5. Spot-check 5 accounts for correct totals",
     "50 concurrent users, 60s",
     "Zero corruption, < 1% errors, cart totals accurate", "", "", "Medium", "", ""],

    ["PT-015", "Vite HMR Speed", "Hot Module Replacement within 500ms",
     "Verify Vite HMR propagates changes almost instantly",
     "Dev server running",
     "1. Open /dashboard\n2. Edit Header.jsx text\n3. Save file\n4. Measure time to visual update\n5. Repeat 5 times, calculate average",
     "Small text change",
     "Average HMR update < 500ms, no full page reload", "", "", "Low", "", ""],

    ["PT-016", "Pagination Last Page Performance", "Last page loads as fast as first",
     "Verify no degradation when requesting last page of paginated catalog",
     "500+ medicines",
     "1. GET page=1&limit=12 -> record time\n2. GET page=42 (last) -> record time\n3. Compare\n4. Verify < 100ms difference",
     "Page 1 vs Page 42",
     "Last page within 100ms of first page, no scan-all penalty", "", "", "Medium", "", ""],

    ["PT-017", "Inventory CSV Upload", "1000-row CSV upload < 10 seconds",
     "Measure inventory CSV upload processing including parsing, validation, DB writes",
     "Salesperson logged in, 1000-row CSV",
     "1. /salesperson/inventory -> select branch\n2. Upload 1000-row CSV\n3. Timer: upload click to success message\n4. Record total time\n5. Verify all 1000 items in DB",
     "CSV: 1000 rows, ~200KB",
     "Upload and processing < 10 seconds, all rows persisted", "", "", "Medium", "", ""],

    ["PT-018", "Session Store Performance", "200 active sessions handled smoothly",
     "Verify MongoDB-backed session store performs under realistic load",
     "Server with connect-mongo session store",
     "1. Create 200 concurrent sessions\n2. Each makes 3 API calls\n3. Measure avg session lookup time\n4. Verify no session conflicts\n5. Monitor sessions collection",
     "200 concurrent sessions",
     "Session lookups < 50ms avg, zero conflicts", "", "", "Low", "", ""],

    ["PT-019", "BookAppointment Page Load", "Appointment booking page loads within 2 seconds",
     "Measure BookAppointment.jsx render time including doctor list and available slots API calls",
     "10+ approved doctors with slots, customer logged in",
     "1. DevTools Performance\n2. Navigate to /appointments/book\n3. Measure time until all doctor cards render\n4. Verify slot availability loads per doctor on selection\n5. Record total interactive time",
     "10+ doctors with slots",
     "Doctor list renders within 2 seconds, slot loading < 1s per doctor", "", "", "Medium", "", ""],

    ["PT-020", "Admin Analytics Filter Speed", "Filtering analytics by date range responds < 1 second",
     "Measure time for analytics charts to re-render when date range filter changes",
     "Admin logged in, analytics page open",
     "1. Navigate to RevenueAnalytics\n2. Set date filter: Last 30 days -> record re-render time\n3. Set filter: Last 90 days -> record time\n4. Set filter: Last 12 months -> record time\n5. Verify charts fully re-render each time",
     "Date ranges: 30d, 90d, 12m",
     "Chart re-render < 1 second for all date ranges", "", "", "Medium", "", ""],

    ["PT-021", "Checkout Page Total Recalculation", "Cart total recalculates within 200ms on coupon apply",
     "Measure the time for Checkout.jsx to recalculate and display updated totals after coupon application",
     "Customer on checkout page with 5+ items in cart",
     "1. Navigate to /checkout with 5 items\n2. Enter coupon code\n3. Click 'Apply'\n4. Measure time from click to total update on screen\n5. Verify discount amount displayed correctly",
     "Cart with 5 items, valid coupon",
     "Total recalculation and display < 200ms after coupon applied", "", "", "Low", "", ""],

    ["PT-022", "First Contentful Paint - All Portals", "FCP < 1.5s for all portal landing pages",
     "Measure First Contentful Paint across all 4 portal landing pages",
     "All portal accounts available, dev server running",
     "1. Lighthouse audit on /customer landing -> record FCP\n2. Lighthouse on /admin/login -> record FCP\n3. Lighthouse on /doctor/login -> record FCP\n4. Lighthouse on /salesperson/login -> record FCP\n5. Compare all values against 1.5s threshold",
     "All 4 portal URLs",
     "FCP < 1.5s for all portal entry points", "", "", "Medium", "", ""],

    ["PT-023", "Admin Customer List Render", "Customer list with 500+ records renders within 3 seconds",
     "Measure AdminCustomerList.jsx render time with large dataset and pagination",
     "500+ customers in database, admin logged in",
     "1. Navigate to /admin/customers\n2. Measure time to first page render (10-20 items)\n3. Click 'Next' page -> measure render time\n4. Search by name -> measure filter response time\n5. Verify no lag or freezing",
     "500+ customer records",
     "First page < 3s, pagination < 1s, search < 500ms", "", "", "Medium", "", ""],

    ["PT-024", "Doctor ConsultationHistory Load", "Consultation history with 100+ records loads efficiently",
     "Measure ConsultationHistory.jsx render time with large history dataset",
     "Doctor with 100+ past consultations",
     "1. Login as Doctor\n2. Navigate to /doctor/consultations\n3. Measure time until consultation list renders\n4. Scroll through list -> verify smooth scrolling\n5. Apply status filter -> measure re-render",
     "100+ consultation records",
     "List renders < 2s, smooth scrolling, filter response < 500ms", "", "", "Low", "", ""],

    ["PT-025", "BranchStatistics Chart Rendering", "Branch statistics with 10+ branches loads under 5 seconds",
     "Measure BranchStatistics.jsx render time including all comparison charts",
     "Admin logged in, 10+ branches with performance data",
     "1. Navigate to /admin/branches/statistics\n2. Measure time until all branch comparison charts render\n3. Verify no chart shows loading state after 5s\n4. Interact with chart (hover) -> verify tooltip response is instant",
     "10+ branches with data",
     "All branch charts rendered within 5 seconds, interactive tooltips responsive", "", "", "Low", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  6. SECURITY TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

security_tests = [
    ["SEC-001", "Session Validation", "Expired session cookie rejected",
     "Verify Express session middleware rejects expired/tampered session cookies",
     "Session created and manually expired",
     "1. Login for valid cookie\n2. Modify cookie expiry to past\n3. GET /api/customer/cart -> verify 401\n4. No data leakage in error",
     "Expired session cookie",
     "401 with generic 'Unauthorized', no sensitive data exposed", "", "", "High", "", ""],

    ["SEC-002", "NoSQL Injection - Login", "MongoDB injection in email field sanitized",
     "Attempt NoSQL injection via login email field",
     "Login page and API accessible",
     "1. Email: { \"$gt\": \"\" }, Password: anything\n2. Submit login\n3. Verify fails\n4. Check logs for handling",
     "Email: {\"$gt\":\"\"}, Password: test",
     "Login fails with 'Invalid Credentials', no database breach", "", "", "High", "", ""],

    ["SEC-003", "XSS - Medicine Search", "Script tags in search sanitized",
     "Inject malicious script via medicine search bar",
     "Medicine catalog loaded",
     "1. Search: <script>alert('XSS')</script>\n2. Submit\n3. No alert popup\n4. DOM shows escaped text",
     "Search: <script>alert('XSS')</script>",
     "No script execution, input escaped or filtered", "", "", "High", "", ""],

    ["SEC-004", "CORS Policy", "Cross-origin requests from unauthorized domains blocked",
     "Verify Express CORS blocks requests from non-allowed origins",
     "Server with CORS configured",
     "1. From evil.com, fetch /api/customer/medicines\n2. Verify CORS error\n3. From localhost:5173, same request -> succeeds",
     "Origin: http://evil.com",
     "Cross-origin blocked; same-origin succeeds", "", "", "High", "", ""],

    ["SEC-005", "Brute Force Protection", "Rate limiter blocks after failed login attempts",
     "Verify authRoutesLimiter blocks excessive login attempts",
     "Rate limiter on auth routes",
     "1. 15 consecutive failed logins\n2. Verify 429 Too Many Requests\n3. Wait for cooldown\n4. Next request succeeds",
     "15 rapid failed attempts",
     "429 after limit, access restored after cooldown", "", "", "High", "", ""],

    ["SEC-006", "Unauthorized Route Access", "Customer cannot access admin endpoints",
     "Verify role-based middleware prevents cross-portal API access",
     "Customer session active",
     "1. Login as Customer\n2. GET /api/admin/dashboard -> 403\n3. POST /api/admin/staff -> 403\n4. Customer routes still work",
     "Customer auth cookie",
     "Admin routes return 403 for customer sessions", "", "", "High", "", ""],

    ["SEC-007", "Password Hash Security", "Passwords stored as bcrypt hashes",
     "Verify passwords are bcrypt hashed in MongoDB, never plaintext",
     "Account registration completed",
     "1. Register with password 'Test12345'\n2. Query MongoDB for customer\n3. Verify passwordHash starts with '$2a$' or '$2b$'\n4. No 'password' field exists",
     "Password: Test12345",
     "DB stores bcrypt hash, no plaintext password field", "", "", "High", "", ""],

    ["SEC-008", "Environment Variable Exposure", "API keys not in frontend bundle",
     "Verify sensitive env vars not included in Vite build output",
     "Production build generated",
     "1. npm run build\n2. Search dist/ for 'mongodb+srv'\n3. Search for 'SESSION_SECRET'\n4. Search for 'GOOGLE_CLIENT_SECRET'\n5. None found",
     "N/A",
     "No server-side secrets in client-side build", "", "", "High", "", ""],

    ["SEC-009", "HTTP Security Headers", "Response includes security headers",
     "Verify Express sends security headers",
     "Server running",
     "1. Make API request\n2. Check X-Frame-Options\n3. Check X-Content-Type-Options: nosniff\n4. Check Strict-Transport-Security",
     "N/A",
     "All security headers present", "", "", "Medium", "", ""],

    ["SEC-010", "Input Length Limits", "Long inputs rejected by Joi validation",
     "Verify Joi DTOs reject excessively long strings",
     "API accessible",
     "1. POST register with 10000 char fullName\n2. Verify 400 from Joi\n3. Server doesn't crash\n4. Error references max length",
     "fullName: 'A' x 10000",
     "400 Validation Error: fullName max length exceeded", "", "", "Medium", "", ""],

    ["SEC-011", "Session Cookie Tampering", "Modified session ID rejected",
     "Manually alter session ID and verify rejection",
     "Valid session via login",
     "1. Login -> copy connect.sid\n2. Modify 3-4 characters\n3. GET /api/customer/cart with tampered cookie\n4. Verify 401\n5. No partial data returned",
     "Tampered connect.sid",
     "Server rejects tampered cookie, 401, no data leakage", "", "", "High", "", ""],

    ["SEC-012", "IDOR - Cross-Customer Order Access", "Customer cannot view another's order",
     "Verify customer cannot access another customer's order by manipulating ID",
     "Two customers, each with orders",
     "1. Login as Customer A\n2. GET /api/customer/orders/:customerB_orderId\n3. Verify 403 or 404\n4. No data leaked",
     "Customer A session, Customer B's order ID",
     "403/404 returned, zero data from other customer's order", "", "", "High", "", ""],

    ["SEC-013", "XSS via Profile Fields", "Script in fullName sanitized on render",
     "Inject script via profile name, verify it doesn't execute anywhere",
     "Customer account exists",
     "1. Edit fullName to: '<img src=x onerror=alert(1)>'\n2. Save\n3. Refresh -> no alert\n4. Admin view -> name escaped",
     "fullName: <img src=x onerror=alert(1)>",
     "Script never executes, name HTML-escaped in all contexts", "", "", "High", "", ""],

    ["SEC-014", "NoSQL Injection - Medicine Search", "Injection payloads in search neutralized",
     "Attempt various NoSQL injection patterns in medicine search",
     "Medicine catalog accessible",
     "1. Search: { '$regex': '.*' } -> not all medicines returned\n2. Search: { '$where': 'sleep(5000)' } -> no delay\n3. Search: { '$ne': null } -> normal behavior",
     "Various NoSQL payloads",
     "All injections neutralized, normal search behavior preserved", "", "", "High", "", ""],

    ["SEC-015", "File Upload Validation", "Inventory rejects non-CSV and malicious files",
     "Verify multer middleware rejects dangerous file types",
     "Salesperson logged in, test files prepared",
     "1. Upload .exe -> rejected\n2. Upload .php -> rejected\n3. Upload CSV with formula injection (=CMD()) -> sanitized\n4. Upload 50MB CSV -> size limit rejection\n5. Upload valid .csv -> accepted",
     "test.exe, test.php, malicious.csv, large.csv, valid.csv",
     "Only valid CSV accepted; dangerous files rejected", "", "", "High", "", ""],

    ["SEC-016", "Role Escalation via API", "User cannot escalate role via profile update",
     "Verify customer cannot change role to 'admin' via PUT request",
     "Customer session active",
     "1. PUT /api/customer/profile with { role: 'admin' }\n2. Verify role NOT updated\n3. Refresh session -> still 'customer'\n4. /api/admin/dashboard -> still 403",
     "Payload: { fullName: 'Test', role: 'admin' }",
     "Server ignores role field, customer role unchanged", "", "", "High", "", ""],

    ["SEC-017", "Session Fixation Prevention", "Session ID regenerated after login",
     "Verify session ID changes after login to prevent fixation",
     "Login page accessible",
     "1. Note pre-login connect.sid\n2. Login successfully\n3. Note new connect.sid\n4. Verify ID changed\n5. Pre-login session invalidated",
     "Pre-login and post-login cookie values",
     "Session ID regenerated after login, old session invalid", "", "", "Medium", "", ""],

    ["SEC-018", "Error Response Safety", "Errors don't leak stack traces or DB info",
     "Verify error responses never include internal details",
     "Various error scenarios",
     "1. Invalid JSON body -> check error response\n2. Non-existent route -> check 404\n3. Invalid ObjectId -> check 500\n4. Verify no 'stack', 'node_modules', or 'mongodb+srv' in any response",
     "Various malformed requests",
     "All errors return clean messages, no internal details", "", "", "Medium", "", ""],

    ["SEC-019", "Verification Token Expiry", "Expired verification token rejected",
     "Verify that expired email verification tokens are rejected and users prompted to request a new link",
     "Customer with expired verificationTokenExpiresAt",
     "1. Register a new customer\n2. Manually set verificationTokenExpiresAt to past date in DB\n3. Click verification link\n4. Verify error: 'Verification link has expired'\n5. Verify user prompted to request new link\n6. Verify account remains unverified",
     "Expired verification token",
     "Expired token rejected, user prompted for new link, account stays unverified", "", "", "Medium", "", ""],

    ["SEC-020", "Password Reset Token Expiry", "Expired reset token rejected gracefully",
     "Verify that expired password reset tokens are rejected with helpful message",
     "Customer with expired resetPasswordExpiresAt",
     "1. Request password reset\n2. Manually expire resetPasswordExpiresAt in DB\n3. Click reset link\n4. Enter new password\n5. Verify error: 'Reset link has expired'\n6. Verify old password still works",
     "Expired reset token",
     "Expired reset link rejected, old password unchanged, user can request new link", "", "", "Medium", "", ""],

    ["SEC-021", "IDOR - Doctor Accessing Other Doctor's Patients", "Doctor cannot view another doctor's consultations",
     "Verify one doctor cannot access consultation data belonging to a different doctor",
     "Two approved doctor accounts, each with consultations",
     "1. Login as Doctor A\n2. GET /api/doctor/consultations/:doctorB_consultationId\n3. Verify 403 Forbidden\n4. Verify no patient data leaked in response\n5. Login as Doctor B -> same consultation -> 200 OK",
     "Doctor A session, Doctor B's consultation ID",
     "Doctor A gets 403, Doctor B gets 200, no cross-doctor data access", "", "", "High", "", ""],

    ["SEC-022", "Account Status Enforcement", "Suspended/blocked accounts cannot login",
     "Verify that customers with account_status set to 'suspended/freezed' or 'blocked/removed' are denied login",
     "Customer accounts with different statuses",
     "1. Set Customer A account_status: 'suspended/freezed' in DB\n2. Attempt login as Customer A -> verify rejected with message\n3. Set Customer B: 'blocked/removed'\n4. Attempt login -> verify rejected\n5. Active account -> login succeeds",
     "Suspended and blocked account credentials",
     "Suspended/blocked accounts denied login with appropriate messages", "", "", "High", "", ""],

    ["SEC-023", "Coupon Code Bruteforce Prevention", "Rate limiting prevents coupon code guessing",
     "Verify that rapid coupon code validation attempts are rate-limited to prevent brute-force discovery",
     "Checkout page accessible, rate limiter configured",
     "1. Send 20 rapid POST requests applying random coupon codes\n2. Verify rate limit triggers (429)\n3. Wait for cooldown\n4. Verify valid coupon code still works after cooldown",
     "20 rapid coupon validation requests",
     "Rate limit activates after threshold, valid code works after cooldown", "", "", "Medium", "", ""],

    ["SEC-024", "Admin Permission Boundary", "Branch-admin cannot access super-admin features",
     "Verify that branch-admin role cannot perform super-admin-only operations like managing other admins",
     "Branch-admin and super-admin accounts exist",
     "1. Login as branch-admin\n2. Try GET /api/admin/staff/admins -> verify 403 (requires manage_admins permission)\n3. Try DELETE /api/admin/branches/:id -> verify 403 if not permitted\n4. Try GET /api/admin/branches (permitted) -> verify 200",
     "Branch-admin credentials",
     "Branch-admin restricted to assigned permissions, super-admin operations blocked", "", "", "High", "", ""],

    ["SEC-025", "CSRF Protection", "Cross-site request forgery attempts blocked",
     "Verify that API endpoints are protected against CSRF attacks via session cookie validation",
     "Customer session active in browser",
     "1. Create a malicious HTML page on different origin\n2. Include form that POSTs to /api/customer/cart/items\n3. Open the malicious page while session active\n4. Verify CORS + SameSite cookie policy blocks the request\n5. Verify no cart modification occurs",
     "Malicious cross-origin form submission",
     "CSRF attempt blocked by CORS and SameSite cookie policy", "", "", "Medium", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  7. USABILITY TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

usability_tests = [
    ["USE-001", "Navigation Clarity", "Main navigation intuitive and accessible",
     "Verify users can easily find and navigate all major sections",
     "Customer logged in, on /dashboard",
     "1. Identify all nav links (Dashboard, Medicines, Cart, Appointments, Orders, Profile)\n2. Click each\n3. Verify correct page loads\n4. Verify active state visible\n5. Verify breadcrumbs (if applicable)",
     "N/A",
     "All links work, active state visible, user knows their location", "", "", "High", "", ""],

    ["USE-002", "Form Error Messages", "Validation errors clear and specific per field",
     "Verify form validation messages appear below their fields with red styling",
     "On /register page",
     "1. Submit empty form\n2. Verify each field shows specific error\n3. Errors are red with icon\n4. Type in field -> error clears\n5. Error text is user-friendly",
     "Empty form submission",
     "Per-field errors with red borders, errors clear on input", "", "", "High", "", ""],

    ["USE-003", "Loading States & Feedback", "All async operations show loading indicators",
     "Verify loading spinners/skeletons during API calls",
     "User logged in",
     "1. /medicines -> spinner before data\n2. 'Add to Cart' -> 'Adding...' state\n3. /cart -> spinner while loading\n4. /appointments -> loading state\n5. /dashboard -> cards show spinners",
     "N/A",
     "Every page shows loading feedback; no blank states during loading", "", "", "High", "", ""],

    ["USE-004", "Empty States", "Empty data shows helpful messages",
     "Verify pages show meaningful empty state messages",
     "Customer with no orders/appointments/cart",
     "1. /orders -> 'No orders found' with icon\n2. /cart -> 'Empty' with 'Browse Medicines' link\n3. /appointments -> 'No appointments'\n4. /notifications -> 'No notifications'",
     "N/A",
     "All empty states show descriptive messages with call-to-action links", "", "", "Medium", "", ""],

    ["USE-005", "Mobile Responsive Layout", "App usable on mobile viewport",
     "Verify pages adapt to 375px mobile viewport",
     "DevTools mobile emulation",
     "1. 375x812 viewport\n2. /dashboard -> vertical stack\n3. /medicines -> 1 column grid\n4. Hamburger menu works\n5. /cart -> full-width checkout button\n6. No horizontal scrollbar",
     "Viewport: 375x812",
     "All pages stack properly, no overflow, touch targets >= 44px", "", "", "Medium", "", ""],

    ["USE-006", "Search UX", "Medicine search provides immediate feedback",
     "Verify search bar provides responsive, helpful feedback",
     "On /medicines with 50+ medicines",
     "1. Type in search bar\n2. Verify trigger after 2-3 chars\n3. 'Searching...' indicator\n4. Non-existent medicine -> 'No medicines found'\n5. Clear search -> full catalog",
     "Search: 'Nonexistent Drug'",
     "Search responsive, feedback during search, no-results handled", "", "", "Medium", "", ""],

    ["USE-007", "Action Confirmation Dialogs", "Destructive actions require confirmation",
     "Verify irreversible actions show confirmation dialogs",
     "User with cart items and pending appointment",
     "1. /cart -> Remove item -> confirm dialog\n2. Cancel -> item stays\n3. OK -> item removed\n4. /appointments -> Cancel -> modal with reason field\n5. /cart -> Clear Cart -> confirmation",
     "N/A",
     "All destructive actions require confirmation before executing", "", "", "Medium", "", ""],

    ["USE-008", "Error Recovery", "Error states guide user toward resolution",
     "Verify API errors show user-friendly messages with guidance",
     "Various error scenarios available",
     "1. 409 (duplicate email) -> 'Email already exists'\n2. 401 (login fail) -> 'Invalid Credentials'\n3. Network failure -> 'Network error. Please try again.'\n4. No raw error codes or stack traces",
     "Various error scenarios",
     "Human-readable messages with clear next steps, no technical details", "", "", "High", "", ""],

    ["USE-009", "Accessibility Basics", "Interactive elements have accessible labels",
     "Verify WCAG accessibility basics: labels, alt text, focus order",
     "Application pages loaded",
     "1. Lighthouse accessibility audit on /dashboard\n2. Check inputs have labels/aria-label\n3. Check images have alt\n4. Check buttons have text\n5. Tab key follows visual layout",
     "N/A",
     "Accessibility score >= 80, proper labels on all elements", "", "", "Low", "", ""],

    ["USE-010", "Consistent Visual Design", "UI components consistent across portals",
     "Verify design system consistency across all 4 portals",
     "Access to all portals",
     "1. Compare button styles across portals\n2. Verify input fields use .input-field class\n3. Verify badge color coding consistent\n4. Card border-radius, shadow, spacing same\n5. Font sizes and weights consistent",
     "N/A",
     "All portals share consistent design language", "", "", "Low", "", ""],

    ["USE-011", "Keyboard Navigation", "All major flows completable via keyboard",
     "Verify user can navigate, fill forms, submit using keyboard only",
     "On /login page",
     "1. Tab -> Email input\n2. Type email -> Tab -> Password\n3. Enter -> form submits\n4. Tab through nav items -> Enter to navigate\n5. Modals close with Escape",
     "Keyboard only (no mouse)",
     "All forms, navigation, actions fully keyboard-accessible", "", "", "Medium", "", ""],

    ["USE-012", "Session Expired Mid-Action", "Session expiry shows helpful redirect",
     "Verify session expiry during form filling shows clear message",
     "User logged in, session expiring",
     "1. Login -> /cart\n2. Let session expire\n3. Click 'Checkout'\n4. Verify 'Session expired' toast/modal\n5. Redirect to /login\n6. After login, redirect back to /cart",
     "Expired session during cart action",
     "Clear message, redirect to login, return to previous page", "", "", "High", "", ""],

    ["USE-013", "Toast Notification Positioning", "Toasts non-blocking and auto-dismiss",
     "Verify toast notifications consistent position, non-blocking, auto-dismiss",
     "Actions that trigger toasts",
     "1. Add to cart -> success toast (top-right)\n2. Toast doesn't block nav\n3. Wait 3-5s -> auto-dismisses\n4. Trigger error -> error toast same position\n5. Multiple toasts -> stack properly",
     "Cart/auth actions",
     "Toasts at top-right, auto-dismiss, stack without overlap", "", "", "Low", "", ""],

    ["USE-014", "Admin Table Pagination & Sorting", "Data tables support pagination, sorting, search",
     "Verify admin tables have functional controls",
     "Admin logged in, 50+ customer records",
     "1. /admin/customers -> shows 10-20 per page\n2. 'Next' -> page 2 loads\n3. Click 'Name' header -> alphabetical sort\n4. Click again -> reverse sort\n5. Search box -> real-time filter",
     "50+ customer records",
     "Pagination, sorting, search all functional", "", "", "Medium", "", ""],

    ["USE-015", "Color Contrast Compliance", "Text meets WCAG 2.1 AA contrast ratios",
     "Verify 4.5:1 minimum contrast ratio for all text",
     "Application loaded in Chrome",
     "1. Lighthouse accessibility audit\n2. Check body text on white backgrounds\n3. Check text on colored badges\n4. Check placeholder text\n5. Check sidebar text on dark background",
     "N/A",
     "All text meets WCAG 2.1 AA contrast ratio 4.5:1", "", "", "Low", "", ""],

    ["USE-016", "Doctor Empty Slot Calendar", "Empty calendar shows 'No slots' with create CTA",
     "Verify empty slot state shows helpful message with action button",
     "Doctor with zero slots",
     "1. Login as Doctor -> /doctor/slots\n2. Verify 'No availability slots configured'\n3. 'Create Your First Slot' button visible\n4. Click CTA -> slot creation form\n5. Create slot -> appears on calendar",
     "Doctor with 0 slots",
     "Empty state with CTA, smooth transition to slot creation", "", "", "Medium", "", ""],

    ["USE-017", "Form Data Persistence Warning", "Unsaved form warns before navigation",
     "Verify navigating away from partially filled form shows warning",
     "On /register page",
     "1. Fill Full Name and Email only\n2. Click navigation link\n3. Verify 'unsaved changes' dialog\n4. 'Stay' -> data preserved\n5. 'Leave' -> navigates away",
     "Partially filled form",
     "Warning appears, 'Stay' preserves data, 'Leave' proceeds", "", "", "Medium", "", ""],

    ["USE-018", "Responsive Admin Sidebar", "Sidebar collapses on tablet viewport",
     "Verify sidebar transitions at breakpoints",
     "Admin logged in, DevTools available",
     "1. 1920x1080 -> full sidebar\n2. 1024x768 -> icons-only\n3. 768x1024 -> hamburger menu\n4. Click hamburger -> overlay\n5. Click outside -> closes",
     "Viewports: 1920, 1024, 768px",
     "Smooth transitions full/collapsed/hamburger at breakpoints", "", "", "Medium", "", ""],

    ["USE-019", "Medicine Detail Image Gallery", "Medicine images can be browsed and zoomed",
     "Verify MedicineDetail.jsx handles multiple product images with navigation",
     "Medicine with 3+ images (img_urls array)",
     "1. Navigate to /medicines/:id with 3 images\n2. Verify primary image displayed large\n3. Verify thumbnail navigation below\n4. Click thumbnail -> primary image changes\n5. Verify image loads without broken placeholder\n6. Test medicine with 0 images -> fallback image shown",
     "Medicine with 3 images, medicine with 0 images",
     "Image gallery navigable, thumbnails work, fallback for missing images", "", "", "Medium", "", ""],

    ["USE-020", "Cart Quantity Adjustment UX", "Cart quantity controls intuitive and responsive",
     "Verify Cart.jsx quantity increment/decrement controls are easy to use with immediate total updates",
     "Customer with items in cart",
     "1. Navigate to /cart\n2. Click '+' button on item -> quantity increases by 1\n3. Verify subtotal updates immediately\n4. Click '-' button -> quantity decreases\n5. At quantity 1, '-' is disabled or shows remove prompt\n6. Verify cart total (bottom) updates with each change",
     "Cart with 2 items",
     "Quantity controls responsive, totals update instantly, min boundary handled", "", "", "Medium", "", ""],

    ["USE-021", "Checkout Address Form", "Checkout address form is clear with proper field labels",
     "Verify Checkout.jsx address form has labeled fields, proper tab order, and error handling",
     "Customer on checkout page",
     "1. Navigate to /checkout\n2. Verify address fields: Street, City, State, Zip Code, Country\n3. Submit without filling -> verify per-field errors\n4. Tab order follows visual layout (top-to-bottom, left-to-right)\n5. Fill all fields -> verify form valid\n6. Verify Map component shows location (if applicable)",
     "Empty address form",
     "Labels clear, tab order logical, per-field errors, map integration works", "", "", "Medium", "", ""],

    ["USE-022", "Doctor Onboarding Progress Indication", "Doctor sees clear progress through onboarding steps",
     "Verify DoctorOnboarding.jsx shows step indicator (Register -> Profile -> Application -> Status)",
     "Doctor in onboarding flow",
     "1. Login as new doctor\n2. Verify step indicator shows current step highlighted\n3. Complete profile -> step advances\n4. Submit application -> status step active\n5. Verify completed steps show checkmark\n6. Verify non-accessible steps are grayed out",
     "Doctor at each onboarding stage",
     "Clear step progression, current step highlighted, completed steps marked", "", "", "Medium", "", ""],

    ["USE-023", "Admin Task Management Filters", "TaskManagement page supports status and priority filters",
     "Verify TaskManagement.jsx has working filter controls for salesperson tasks",
     "Admin logged in, 20+ tasks with different statuses and priorities",
     "1. Navigate to /admin/tasks\n2. Filter by status: 'pending' -> verify only pending shown\n3. Filter by priority: 'urgent' -> verify only urgent shown\n4. Combine filters -> verify intersection\n5. Clear all filters -> all tasks shown\n6. Verify task count updates with filters",
     "20+ tasks with mixed statuses/priorities",
     "Filters work individually and combined, counts update correctly", "", "", "Medium", "", ""],

    ["USE-024", "AddEditReminderModal Usability", "Reminder modal has clear inputs and feedback",
     "Verify the AddEditReminderModal.jsx provides clear inputs for medicine selection, frequency, and time",
     "Customer on /reminders page",
     "1. Click 'Add Reminder'\n2. Verify modal opens with medicine dropdown, frequency selector, time picker\n3. Verify medicine dropdown is searchable\n4. Select frequency: 'daily'\n5. Set time: '08:00'\n6. Submit -> verify success feedback\n7. Verify modal closes automatically",
     "Medicine: Panadol, daily, 08:00",
     "Modal inputs clear, medicine searchable, success feedback, auto-close", "", "", "Low", "", ""],

    ["USE-025", "Error Page (404) UX", "Non-existent routes show helpful 404 page",
     "Verify navigating to invalid routes shows a user-friendly 404 page with navigation options",
     "Application running",
     "1. Navigate to /nonexistent-page\n2. Verify 404 page renders (not blank white screen)\n3. Verify message: 'Page not found'\n4. Verify 'Go to Dashboard' or 'Go Home' link exists\n5. Click link -> navigate back to valid page",
     "URL: /nonexistent-page",
     "Friendly 404 page with clear message and navigation back to valid page", "", "", "Low", "", ""],
]

# ═══════════════════════════════════════════════════════════════════════════════
#  8. COMPATIBILITY TESTING (25 test cases)
# ═══════════════════════════════════════════════════════════════════════════════

compatibility_tests = [
    ["COM-001", "Chrome Browser", "Full functionality in Google Chrome",
     "Verify all features work in Chrome (latest version)",
     "Chrome v120+",
     "1. Open Philbox in Chrome\n2. Register account\n3. Navigate all pages\n4. Cart operations\n5. Appointment booking\n6. Verify CSS\n7. Check console for errors",
     "Chrome v120+",
     "100% feature parity, no issues", "", "", "High", "", ""],

    ["COM-002", "Firefox Browser", "Full functionality in Mozilla Firefox",
     "Verify all features work in Firefox with proper CSS",
     "Firefox v120+",
     "1. Open in Firefox\n2. Landing page rendering\n3. /dashboard\n4. /medicines with filters\n5. Cart and checkout\n6. Verify CSS identical to Chrome\n7. Check console",
     "Firefox v120+",
     "All features work, CSS correct, no JS errors", "", "", "High", "", ""],

    ["COM-003", "Microsoft Edge", "Full functionality in Edge Chromium",
     "Verify Philbox works in Edge browser",
     "Edge installed",
     "1. Open in Edge\n2. Registration flow\n3. Medicine search and cart\n4. Appointment management\n5. OAuth Google popup works\n6. Check console",
     "Edge Chromium v120+",
     "Full functionality, OAuth popup handled correctly", "", "", "Medium", "", ""],

    ["COM-004", "Safari Browser", "Core functionality in Safari",
     "Verify core features in Safari, noting WebKit differences",
     "Safari available (macOS/iOS)",
     "1. Open in Safari\n2. Session cookies work (SameSite)\n3. Main user flows\n4. CSS flexbox/grid\n5. Date input fields\n6. JS compatibility",
     "Safari v17+",
     "Core features functional, CSS acceptable, dates work", "", "", "Medium", "", ""],

    ["COM-005", "Mobile Chrome (Android)", "App responsive on Android phones",
     "Verify responsive rendering on mobile Chrome with touch interactions",
     "Android device or emulation",
     "1. Open on mobile Chrome\n2. Hamburger menu works\n3. /medicines -> 1-2 column grid\n4. Touch interactions work\n5. Form input with virtual keyboard\n6. Cart operations on mobile",
     "Android (or emulated 360x640)",
     "Fully responsive, touch works, keyboard doesn't break layout", "", "", "Medium", "", ""],

    ["COM-006", "Mobile Safari (iOS)", "App responsive on iPhone Safari",
     "Verify Philbox works on iOS Safari with proper viewport",
     "iOS device or simulator",
     "1. Open on iOS Safari\n2. Viewport meta works\n3. Navigation/transitions\n4. Date picker on iOS\n5. Fixed header doesn't bounce\n6. OAuth flow on iOS",
     "iPhone (or emulated 375x812)",
     "Functional on iOS Safari, no viewport issues, dates usable", "", "", "Medium", "", ""],

    ["COM-007", "Resolution - 1920x1080 (FHD)", "Desktop FHD renders correctly",
     "Verify layout utilizes FHD screens appropriately",
     "Browser at 1920x1080",
     "1. /dashboard -> content centered with max-width\n2. /medicines -> 4-5 column grid\n3. Admin charts fill space\n4. No excessive white space",
     "Resolution: 1920x1080",
     "Layout proportional, centered, grids effective", "", "", "Medium", "", ""],

    ["COM-008", "Resolution - 1366x768 (HD)", "Standard laptop resolution OK",
     "Verify layout works on common laptop resolution",
     "Browser at 1366x768",
     "1. Navigate all portal pages\n2. Sidebar doesn't overlap content\n3. Tables scroll horizontally if needed\n4. Modals fit viewport\n5. Text readable without zoom",
     "Resolution: 1366x768",
     "All content visible, no overflow, modals fit", "", "", "Medium", "", ""],

    ["COM-009", "Resolution - 768x1024 (iPad)", "Tablet portrait renders two-column",
     "Verify responsive breakpoints for tablet viewport",
     "Browser at 768x1024",
     "1. /dashboard -> 2-column cards\n2. Sidebar collapses\n3. /medicines -> 2-3 columns\n4. Forms on /register\n5. Admin dashboard",
     "Resolution: 768x1024",
     "Two-column layout, collapsed sidebar, forms usable", "", "", "Low", "", ""],

    ["COM-010", "Node.js Version", "App runs on Node.js v18 and v20 LTS",
     "Verify Express backend runs on supported Node.js LTS versions",
     "Node.js v18 and v20 available",
     "1. Switch to v18\n2. npm install && npm run dev\n3. Server starts OK\n4. Test API call\n5. Switch to v20\n6. Repeat steps 2-4",
     "Node.js v18.x and v20.x",
     "Server runs cleanly on both, no breaking issues", "", "", "Low", "", ""],

    ["COM-011", "Resolution - 2560x1440 (QHD)", "2K display renders without scaling issues",
     "Verify proper scaling on high-resolution 2K monitors",
     "Browser at 2560x1440",
     "1. 2560x1440 at 100%\n2. /dashboard -> content not tiny\n3. Text readable\n4. Images not pixelated\n5. Charts scale proportionally",
     "Resolution: 2560x1440, Scale: 100%",
     "Content scaled, text crisp, images sharp, layouts proportional", "", "", "Low", "", ""],

    ["COM-012", "Resolution - 360x640 (Small Android)", "Smallest common Android viewport",
     "Verify app works on minimum target mobile viewport",
     "Browser at 360x640",
     "1. /login -> form fits without scroll\n2. /medicines -> single column\n3. Cart items don't overflow\n4. Checkout inputs wide enough\n5. Buttons >= 44px touch target",
     "Resolution: 360x640",
     "All content renders usably, no overflow, adequate targets", "", "", "Medium", "", ""],

    ["COM-013", "Windows OS", "Full functionality on Windows 10 and 11",
     "Verify Philbox runs on Chrome on both Windows versions",
     "Windows 10 and 11 machines",
     "1. Chrome on Windows 10 -> full customer journey\n2. Admin portal with file upload\n3. Repeat on Windows 11\n4. Font rendering correct\n5. Date pickers work on both",
     "OS: Windows 10, Windows 11",
     "Full functionality on both Windows versions", "", "", "Medium", "", ""],

    ["COM-014", "macOS Testing", "Full functionality on macOS",
     "Verify Philbox works on Safari and Chrome on macOS",
     "macOS with Safari and Chrome",
     "1. Safari on macOS -> CSS rendering\n2. File upload for inventory\n3. OAuth popup handling\n4. Chrome comparison\n5. Trackpad scroll smooth",
     "macOS Ventura/Sonoma, Safari + Chrome",
     "Full functionality in both browsers, smooth scroll", "", "", "Low", "", ""],

    ["COM-015", "Landscape Mode (Mobile)", "Mobile landscape usable",
     "Verify landscape orientation works without breaks",
     "Mobile device or emulated landscape",
     "1. 812x375 (landscape)\n2. /dashboard -> horizontal layout\n3. /medicines -> 2-3 columns\n4. Modal fits landscape\n5. Flip to portrait -> recovers\n6. No persistent scrollbar",
     "812x375 (landscape), 375x812 (portrait)",
     "Landscape usable, modals fit, smooth orientation transitions", "", "", "Low", "", ""],

    ["COM-016", "MongoDB v6 vs v7", "Backend works with MongoDB 6.x and 7.x",
     "Verify Mongoose schemas and queries work on both versions",
     "MongoDB 6.x and 7.x instances",
     "1. Point backend to MongoDB 6.x\n2. Seed data -> test CRUD\n3. Run analytics aggregations\n4. Switch to MongoDB 7.x\n5. Repeat -> compare results",
     "MongoDB 6.x and 7.x",
     "All operations identical on both versions", "", "", "Low", "", ""],

    ["COM-017", "Browser Dark Mode", "App handles OS dark mode",
     "Verify application handles system dark mode without breaking",
     "OS with dark mode toggle",
     "1. Enable OS dark mode\n2. Open Philbox\n3. Check UI: adapts or maintains own theme\n4. All text readable\n5. Input field borders visible\n6. Disable dark mode -> normal",
     "OS Dark Mode: ON then OFF",
     "No UI breakage or unreadable text with dark mode", "", "", "Low", "", ""],

    ["COM-018", "High Latency Network", "App functions with 500ms latency",
     "Verify usable on high-latency connections",
     "DevTools throttle or traffic shaper",
     "1. 500ms RTT latency\n2. Login -> works (slow)\n3. /medicines -> loads\n4. Add to cart -> completes\n5. No timeouts or duplicate submissions",
     "Network: 500ms RTT",
     "All features work (slower), no timeouts or duplicates", "", "", "Medium", "", ""],

    ["COM-019", "Customer Portal - Incognito/Private Mode", "App works in incognito browser mode",
     "Verify session management and cookies work correctly in private browsing mode",
     "Chrome Incognito or Firefox Private Window",
     "1. Open Incognito window\n2. Navigate to Philbox\n3. Register/Login -> verify session works\n4. Navigate through pages\n5. Close incognito -> reopen -> verify session cleared\n6. Google OAuth -> verify popup works in incognito",
     "Incognito/Private browsing mode",
     "Full functionality in incognito, session clears on window close", "", "", "Medium", "", ""],

    ["COM-020", "Admin Portal - Large Screen (4K)", "Admin portal usable on 4K display",
     "Verify admin dashboard and tables are usable on 3840x2160 resolution at 150% scaling",
     "4K display or emulated resolution",
     "1. Set browser to 3840x2160 at 150% scaling\n2. Navigate /admin/dashboard -> verify charts readable\n3. /admin/customers -> verify table columns don't stretch excessively\n4. /admin/branches -> verify cards properly spaced\n5. Modal dialogs centered and sized appropriately",
     "Resolution: 3840x2160 at 150% scale",
     "Admin UI proportional and usable at 4K, nothing overly stretched", "", "", "Low", "", ""],

    ["COM-021", "Doctor Portal - Samsung Browser", "Doctor portal works on Samsung Internet Browser",
     "Verify core doctor portal functionality on Samsung Internet (popular Android alternative browser)",
     "Samsung Internet browser v20+",
     "1. Open doctor portal login\n2. Login with doctor credentials\n3. Navigate to /doctor/appointments\n4. View slot management\n5. Test consultation flow\n6. Check CSS rendering against Chrome",
     "Samsung Internet v20+",
     "Core functionality works, CSS renders acceptably", "", "", "Low", "", ""],

    ["COM-022", "Slow CPU - Budget Device", "App usable on low-end device with slow CPU",
     "Verify application doesn't freeze or become unusable on budget devices with CPU throttling",
     "DevTools CPU throttle 4x slowdown",
     "1. DevTools -> Performance -> CPU: 4x slowdown\n2. Navigate to /dashboard -> verify renders without freeze\n3. Navigate to /medicines -> verify scroll is smooth\n4. Open medicine detail -> verify interaction responsive\n5. Add to cart -> verify action completes within 3s",
     "CPU: 4x slowdown throttle",
     "App usable (slower) without freezing or becoming unresponsive", "", "", "Medium", "", ""],

    ["COM-023", "Multiple Tab Session Consistency", "Same user in multiple tabs has consistent state",
     "Verify opening multiple browser tabs for the same user maintains consistent session state",
     "Customer logged in",
     "1. Open Tab 1: /dashboard\n2. Open Tab 2: /medicines\n3. Tab 2: add to cart -> verify badge updates\n4. Tab 1: refresh -> verify cart count matches\n5. Tab 2: logout -> Tab 1: refresh -> verify logged out",
     "Single user, multiple tabs",
     "Session consistent across tabs, logout propagates", "", "", "Medium", "", ""],

    ["COM-024", "Admin Portal - Firefox on Linux", "Admin portal works on Firefox on Ubuntu/Linux",
     "Verify admin portal CSS and functionality on Firefox running on Linux",
     "Ubuntu/Linux with Firefox",
     "1. Open admin portal in Firefox on Ubuntu\n2. Verify font rendering (Inter/system fonts)\n3. Navigate /admin/dashboard -> charts render\n4. Test file upload for inventory\n5. Verify modal positioning\n6. Check scrollbar styling",
     "Ubuntu 22.04+, Firefox v120+",
     "Full functionality, fonts render correctly, scrollbars styled", "", "", "Low", "", ""],

    ["COM-025", "Offline/Service Worker Behavior", "App shows meaningful offline state (no service worker crash)",
     "Verify the application handles complete offline state without crashing or showing confusing errors",
     "Application loaded, network available for initial load",
     "1. Load /dashboard fully\n2. Set DevTools Network: Offline\n3. Click /medicines -> verify meaningful 'No connection' message\n4. Verify app doesn't crash (no white screen)\n5. Re-enable network -> verify app recovers\n6. Verify any cached content still accessible during offline",
     "Network: Offline after initial load",
     "Offline state shows meaningful message, no crash, recovers on reconnect", "", "", "Low", "", ""],
]


# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

all_sheets_data = [
    ("Unit Testing",          unit_tests),
    ("Integration Testing",   integration_tests),
    ("System Testing",        system_tests),
    ("Acceptance Testing",    acceptance_tests),
    ("Performance Testing",   performance_tests),
    ("Security Testing",      security_tests),
    ("Usability Testing",     usability_tests),
    ("Compatibility Testing", compatibility_tests),
]

summary_rows = []

for idx, (sheet_name, test_data) in enumerate(all_sheets_data):
    if idx == 0:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    theme = THEMES[sheet_name]
    header_fill = PatternFill(start_color=theme[0], end_color=theme[0], fill_type="solid")
    alt_fill = PatternFill(start_color=theme[2], end_color=theme[2], fill_type="solid")
    ws.sheet_properties.tabColor = theme[1]

    # Write header
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Write test data
    data_font = Font(size=10, name="Calibri")
    for row_idx, test_row in enumerate(test_data, 2):
        for col_idx, value in enumerate(test_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Set row heights
    ws.row_dimensions[1].height = 28
    for row_idx in range(2, len(test_data) + 2):
        ws.row_dimensions[row_idx].height = 80

    # Track for summary
    summary_rows.append([sheet_name, len(test_data), 0, 0, len(test_data)])


# ─── SUMMARY SHEET ─────────────────────────────────────────────────────────
ws_summary = wb.create_sheet(title="Summary")
ws_summary.sheet_properties.tabColor = "2C3E50"

summary_headers = [
    "Testing Type", "Total Test Cases", "Passed", "Failed", "Pending"
]

# Title
title_cell = ws_summary.cell(row=1, column=1, value="PHILBOX - Software Testing Summary Report")
title_cell.font = Font(bold=True, size=16, color="2C3E50", name="Calibri")
ws_summary.merge_cells('A1:E1')
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Subtitle
sub_cell = ws_summary.cell(row=2, column=1, value="Project: Philbox | Tech Stack: MERN (MongoDB, Express.js, React.js, Node.js)")
sub_cell.font = Font(size=11, color="7F8C8D", name="Calibri")
ws_summary.merge_cells('A2:E2')
sub_cell.alignment = Alignment(horizontal="center")

# Blank row
ws_summary.row_dimensions[3].height = 10

# Summary table header
sum_header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
sum_header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
for col_idx, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=4, column=col_idx, value=header)
    cell.font = sum_header_font
    cell.fill = sum_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

# Summary data
total_all = 0
alt_sum_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
for row_idx, srow in enumerate(summary_rows, 5):
    data_font_s = Font(size=10, name="Calibri")
    for col_idx, val in enumerate(srow, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font_s
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if row_idx % 2 == 0:
            cell.fill = alt_sum_fill
    total_all += srow[1]

# Total row
total_row = len(summary_rows) + 5
for col_idx, val in enumerate(["TOTAL", total_all, 0, 0, total_all], 1):
    cell = ws_summary.cell(row=total_row, column=col_idx, value=val)
    cell.font = Font(bold=True, size=11, name="Calibri")
    cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

# Notes section
notes_row = total_row + 2
ws_summary.cell(row=notes_row, column=1, value="Notes:").font = Font(bold=True, size=11, name="Calibri")
ws_summary.cell(row=notes_row + 1, column=1, value="* 'Actual Result' and 'Status' columns are left blank -- to be filled during testing execution.").font = Font(size=10, color="7F8C8D", name="Calibri")
ws_summary.cell(row=notes_row + 2, column=1, value="* Priority levels: High = Critical business function, Medium = Important but not blocking, Low = Nice to have.").font = Font(size=10, color="7F8C8D", name="Calibri")
ws_summary.cell(row=notes_row + 3, column=1, value="* Test Case IDs: UT=Unit, IT=Integration, ST=System, AT=Acceptance, PT=Performance, SEC=Security, USE=Usability, COM=Compatibility.").font = Font(size=10, color="7F8C8D", name="Calibri")
ws_summary.merge_cells(f'A{notes_row+1}:E{notes_row+1}')
ws_summary.merge_cells(f'A{notes_row+2}:E{notes_row+2}')
ws_summary.merge_cells(f'A{notes_row+3}:E{notes_row+3}')

# Summary column widths
for col_idx, w in enumerate([30, 18, 12, 12, 12], 1):
    ws_summary.column_dimensions[get_column_letter(col_idx)].width = w

ws_summary.freeze_panes = "A5"


# ─── SAVE ────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\n[OK] Philbox Software Testing Document generated successfully!")
print(f"File: {OUTPUT_PATH}")
print(f"Total Test Cases: {total_all}")
print(f"Sheets: {len(all_sheets_data) + 1} (8 test types + 1 summary)")
